from pathlib import Path

import openpyxl

CHECKLIST_PATH = Path(__file__).resolve().parent / "Form checklist hồ sơ dự án.xlsx"
SHEET_VIAM = "Đề tài - Bánh ăn dặm VIAM 2027"
CORRECT_FILENAME = "Lý lịch khoa học - Trương Hồng Sơn.docx"


def _find_row(ws, code: str) -> int:
    for row in ws.iter_rows(min_row=5, max_col=1):
        if row[0].value == code:
            return row[0].row
    raise ValueError(f"Không tìm thấy mã mục '{code}' trong sheet '{ws.title}'")


def fix_f01_cv_filename() -> None:
    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    ws = wb[SHEET_VIAM]
    row = _find_row(ws, "F01")
    if ws.cell(row=row, column=5).value == CORRECT_FILENAME:
        return
    ws.cell(row=row, column=5, value=CORRECT_FILENAME)
    wb.save(CHECKLIST_PATH)


if __name__ == "__main__":
    fix_f01_cv_filename()
    print(f"Da cap nhat F01 thanh '{CORRECT_FILENAME}'.")
