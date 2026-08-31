# migrate_add_contact_person.py
import copy
import re
from pathlib import Path

import openpyxl

CHECKLIST_PATH = Path(__file__).resolve().parent / "Form checklist hồ sơ dự án.xlsx"
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]
INSERT_AT_ROW = 13
STYLE_SOURCE_ROW = 12


def _fix_status_formula_row_refs(ws) -> None:
    for row_cells in ws.iter_rows(min_row=5):
        code_cell = row_cells[0]
        if not isinstance(code_cell.value, str):
            continue
        formula_cell = row_cells[5]
        if not (isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")):
            continue
        actual_row = code_cell.row
        formula_cell.value = re.sub(
            r"\b([CDE])(\d{1,3})\b",
            lambda m: f"{m.group(1)}{actual_row}",
            formula_cell.value,
        )


def _has_contact_person_field(ws) -> bool:
    for row in ws.iter_rows(min_row=5, max_col=1):
        if row[0].value == "A08":
            return True
    return False


def add_contact_person_field() -> None:
    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    changed = False
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        if _has_contact_person_field(ws):
            continue
        changed = True
        ws.insert_rows(INSERT_AT_ROW)

        for col in range(1, 7):
            src_cell = ws.cell(row=STYLE_SOURCE_ROW, column=col)
            dst_cell = ws.cell(row=INSERT_AT_ROW, column=col)
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)

        ws.cell(row=INSERT_AT_ROW, column=1, value="A08")
        ws.cell(
            row=INSERT_AT_ROW,
            column=2,
            value="Đầu mối liên hệ (họ tên, đơn vị, email, điện thoại - Tùy chọn)",
        )
        ws.cell(
            row=INSERT_AT_ROW,
            column=6,
            value=f'=IF(ISBLANK(C{INSERT_AT_ROW}), "⚪ Tùy chọn (Trống)", "✅ Xong")',
        )

        _fix_status_formula_row_refs(ws)

    if changed:
        wb.save(CHECKLIST_PATH)


if __name__ == "__main__":
    add_contact_person_field()
    print("Da them truong A08 'Dau moi lien he' vao ca 2 sheet.")
