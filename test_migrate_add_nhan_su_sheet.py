# test_migrate_add_nhan_su_sheet.py
import openpyxl

import migrate_add_nhan_su_sheet as migrate


def _build_project_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    row_num = 5
    for code, name, degree, org in rows:
        ws.cell(row=row_num, column=1, value=code)
        ws.cell(row=row_num, column=3, value=name)
        ws.cell(row=row_num, column=4, value=degree)
        ws.cell(row=row_num, column=5, value=org)
        row_num += 1
    return ws


def test_add_nhan_su_sheet_creates_hidden_sheet_with_headers(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_project_sheet(wb, "Đề tài - Bánh ăn dặm VIAM 2027", [("B01", "Trương Hồng Sơn", "TS.BS.", "VIAM")])
    _build_project_sheet(wb, "Đề tài - Mẫu trắng dự án mới", [])
    path = tmp_path / "checklist.xlsx"
    wb.save(path)

    migrate.add_nhan_su_sheet(path)

    result = openpyxl.load_workbook(path)
    assert migrate.NHAN_SU_SHEET_NAME in result.sheetnames
    ws = result[migrate.NHAN_SU_SHEET_NAME]
    assert ws.sheet_state == "hidden"
    assert [ws.cell(row=1, column=c).value for c in range(1, 4)] == ["ten", "hoc_ham_hoc_vi", "don_vi"]


def test_add_nhan_su_sheet_seeds_distinct_names_from_both_project_sheets(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_project_sheet(
        wb,
        "Đề tài - Bánh ăn dặm VIAM 2027",
        [("B01", "Trương Hồng Sơn", "TS.BS.", "VIAM"), ("C01", "Nguyễn Công Khẩn", "GS.TS.", "Hội đồng Đạo đức")],
    )
    _build_project_sheet(wb, "Đề tài - Mẫu trắng dự án mới", [("B01", "Trương Hồng Sơn", "TS.BS.", "VIAM")])
    path = tmp_path / "checklist.xlsx"
    wb.save(path)

    migrate.add_nhan_su_sheet(path)

    result = openpyxl.load_workbook(path)
    ws = result[migrate.NHAN_SU_SHEET_NAME]
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert names.count("Trương Hồng Sơn") == 1
    assert "Nguyễn Công Khẩn" in names


def test_add_nhan_su_sheet_is_idempotent(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_project_sheet(wb, "Đề tài - Bánh ăn dặm VIAM 2027", [("B01", "Trương Hồng Sơn", "TS.BS.", "VIAM")])
    _build_project_sheet(wb, "Đề tài - Mẫu trắng dự án mới", [])
    path = tmp_path / "checklist.xlsx"
    wb.save(path)

    migrate.add_nhan_su_sheet(path)
    migrate.add_nhan_su_sheet(path)

    result = openpyxl.load_workbook(path)
    ws = result[migrate.NHAN_SU_SHEET_NAME]
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert names.count("Trương Hồng Sơn") == 1


def test_add_nhan_su_sheet_preserves_manually_added_contact_info(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_project_sheet(wb, "Đề tài - Bánh ăn dặm VIAM 2027", [("B01", "Trương Hồng Sơn", "TS.BS.", "VIAM")])
    _build_project_sheet(wb, "Đề tài - Mẫu trắng dự án mới", [])
    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    migrate.add_nhan_su_sheet(path)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2[migrate.NHAN_SU_SHEET_NAME]
    ws2.cell(row=2, column=4, value="Số 47 Đặng Văn Ngữ, Hà Nội")
    wb2.save(path)

    migrate.add_nhan_su_sheet(path)

    result = openpyxl.load_workbook(path)
    ws = result[migrate.NHAN_SU_SHEET_NAME]
    assert ws.cell(row=2, column=4).value == "Số 47 Đặng Văn Ngữ, Hà Nội"
