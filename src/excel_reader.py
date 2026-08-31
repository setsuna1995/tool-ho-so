import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl

CHECKLIST_FILENAME = "Form checklist hồ sơ dự án.xlsx"

TIMELINE_PATTERN = re.compile(r"(\d{2})/(\d{4}).*?(\d{2})/(\d{4})")


def parse_timeline(text: str) -> Tuple[str, str]:
    """Tách mốc bắt đầu/kết thúc dạng MM/YYYY từ nội dung mốc thời gian (A05)."""
    match = TIMELINE_PATTERN.search(text)
    if not match:
        raise ValueError(
            f"Không đọc được mốc thời gian nghiên cứu (A05) từ nội dung '{text}'. "
            "Vui lòng nhập đúng định dạng có 2 mốc MM/YYYY, ví dụ: "
            "'Tháng 01/2027 đến tháng 12/2027'."
        )
    start_month, start_year, end_month, end_year = match.groups()
    return f"{start_month}/{start_year}", f"{end_month}/{end_year}"


@dataclass
class Person:
    name: str
    degree: str = ""
    org: str = ""


@dataclass
class ExpertCvEntry:
    code: str
    name: str
    role: str


@dataclass
class CommitteeData:
    chair: Person
    reviewers: List[Person] = field(default_factory=list)
    members: List[Person] = field(default_factory=list)
    secretaries: List[Person] = field(default_factory=list)


@dataclass
class ProjectInfo:
    title: str
    research_type: str
    year: int
    host_org: str
    partner_org: Optional[str]
    research_location: Optional[str]
    timeline: str
    head: Person
    co_head: Optional[Person]
    project_secretary: Optional[Person]
    researchers: List[Person]
    ethics_committee: CommitteeData
    proposal_committee: CommitteeData
    acceptance_committee: CommitteeData
    expert_cvs: List[ExpertCvEntry]
    package_id: Optional[str] = None
    common_tokens: dict = field(default_factory=dict)


CODE_ALIASES: dict = {
    "CFG_PACKAGE": ["CFG_PACKAGE", "{{CFG_PACKAGE}}", "BO_HO_SO", "{{BO_HO_SO}}", "A09"],
    "A01": ["A01", "TEN_DE_TAI", "{{TEN_DE_TAI}}"],
    "A02": ["A02", "LOAI_HINH_NC", "{{LOAI_HINH_NC}}", "KIEU_NGHIEN_CUU", "{{KIEU_NGHIEN_CUU}}"],
    "A03": ["A03", "NAM", "{{NAM}}"],
    "A04": ["A04", "DON_VI_CHU_TRI", "{{DON_VI_CHU_TRI}}", "CO_QUAN_CHU_TRI"],
    "A05": ["A05", "THOI_GIAN_BAT_DAU", "{{THOI_GIAN_BAT_DAU}}", "THOI_GIAN_KET_THUC", "{{THOI_GIAN_KET_THUC}}", "THOI_GIAN_TRIEN_KHAI", "{{THOI_GIAN_TRIEN_KHAI}}"],
    "A06": ["A06", "DON_VI_DOI_TAC", "{{DON_VI_DOI_TAC}}", "CO_QUAN_PHOI_HOP"],
    "A07": ["A07", "DIA_DIEM_TRIEN_KHAI", "{{DIA_DIEM_TRIEN_KHAI}}"],
    "A08": ["A08", "DAU_MOI_LIEN_HE", "{{DAU_MOI_LIEN_HE}}"],
    "A09": ["A09", "CFG_PACKAGE", "{{CFG_PACKAGE}}", "BO_HO_SO", "{{BO_HO_SO}}"],
    "B01": ["B01", "CHU_NHIEM_HO_TEN", "{{CHU_NHIEM_HO_TEN}}", "CHU_NHIEM_TEN", "{{CHU_NHIEM_TEN}}", "CHU_NHIEM_DON_VI", "{{CHU_NHIEM_DON_VI}}"],
    "B02": ["B02", "DONG_CHU_NHIEM_HO_TEN", "{{DONG_CHU_NHIEM_HO_TEN}}", "DONG_CHU_NHIEM_TEN", "{{DONG_CHU_NHIEM_TEN}}"],
    "B03": ["B03", "THU_KY_DE_TAI", "{{THU_KY_DE_TAI}}"],
    "B04": ["B04", "B04_1", "NGHIEN_CUU_VIEN_1", "{{NGHIEN_CUU_VIEN_1}}", "DANH_SACH_NGHIEN_CUU_VIEN", "{{DANH_SACH_NGHIEN_CUU_VIEN}}"],
    "B05": ["B05", "B04_2", "NGHIEN_CUU_VIEN_2", "{{NGHIEN_CUU_VIEN_2}}"],
    "B06": ["B06", "B04_3", "NGHIEN_CUU_VIEN_3", "{{NGHIEN_CUU_VIEN_3}}"],
    "B07": ["B07", "B04_4", "NGHIEN_CUU_VIEN_4", "{{NGHIEN_CUU_VIEN_4}}"],
    "B08": ["B08", "B04_5", "NGHIEN_CUU_VIEN_5", "{{NGHIEN_CUU_VIEN_5}}"],
    "B09": ["B09", "B04_6", "NGHIEN_CUU_VIEN_6", "{{NGHIEN_CUU_VIEN_6}}"],
    "B10": ["B10", "B04_7", "NGHIEN_CUU_VIEN_7", "{{NGHIEN_CUU_VIEN_7}}"],
    "B11": ["B11", "B04_8", "NGHIEN_CUU_VIEN_8", "{{NGHIEN_CUU_VIEN_8}}"],
    "B12": ["B12", "B04_9", "NGHIEN_CUU_VIEN_9", "{{NGHIEN_CUU_VIEN_9}}"],
    "B13": ["B13", "B04_10", "NGHIEN_CUU_VIEN_10", "{{NGHIEN_CUU_VIEN_10}}"],
    "B14": ["B14", "B04_11", "NGHIEN_CUU_VIEN_11", "{{NGHIEN_CUU_VIEN_11}}"],
    "B15": ["B15", "B04_12", "NGHIEN_CUU_VIEN_12", "{{NGHIEN_CUU_VIEN_12}}"],
    "B16": ["B16", "B04_13", "NGHIEN_CUU_VIEN_13", "{{NGHIEN_CUU_VIEN_13}}"],
    "B17": ["B17", "B04_14", "NGHIEN_CUU_VIEN_14", "{{NGHIEN_CUU_VIEN_14}}"],
    "B18": ["B18", "B04_15", "NGHIEN_CUU_VIEN_15", "{{NGHIEN_CUU_VIEN_15}}"],
    "B19": ["B19", "B04_16", "NGHIEN_CUU_VIEN_16", "{{NGHIEN_CUU_VIEN_16}}"],
    "B20": ["B20", "B04_17", "NGHIEN_CUU_VIEN_17", "{{NGHIEN_CUU_VIEN_17}}"],
    "B04_1": ["B04_1", "B04", "NGHIEN_CUU_VIEN_1", "{{NGHIEN_CUU_VIEN_1}}"],
    "B04_2": ["B04_2", "B05", "NGHIEN_CUU_VIEN_2", "{{NGHIEN_CUU_VIEN_2}}"],
    "B04_3": ["B04_3", "B06", "NGHIEN_CUU_VIEN_3", "{{NGHIEN_CUU_VIEN_3}}"],
    "B04_4": ["B04_4", "B07", "NGHIEN_CUU_VIEN_4", "{{NGHIEN_CUU_VIEN_4}}"],
    "B04_5": ["B04_5", "B08", "NGHIEN_CUU_VIEN_5", "{{NGHIEN_CUU_VIEN_5}}"],
    "B04_6": ["B04_6", "B09", "NGHIEN_CUU_VIEN_6", "{{NGHIEN_CUU_VIEN_6}}"],
    "B04_7": ["B04_7", "B10", "NGHIEN_CUU_VIEN_7", "{{NGHIEN_CUU_VIEN_7}}"],
    "B04_8": ["B04_8", "B11", "NGHIEN_CUU_VIEN_8", "{{NGHIEN_CUU_VIEN_8}}"],
    "B04_9": ["B04_9", "B12", "NGHIEN_CUU_VIEN_9", "{{NGHIEN_CUU_VIEN_9}}"],
    "B04_10": ["B04_10", "B13", "NGHIEN_CUU_VIEN_10", "{{NGHIEN_CUU_VIEN_10}}"],
    "B04_11": ["B04_11", "B14", "NGHIEN_CUU_VIEN_11", "{{NGHIEN_CUU_VIEN_11}}"],
    "B04_12": ["B04_12", "B15", "NGHIEN_CUU_VIEN_12", "{{NGHIEN_CUU_VIEN_12}}"],
    "B04_13": ["B04_13", "B16", "NGHIEN_CUU_VIEN_13", "{{NGHIEN_CUU_VIEN_13}}"],
    "B04_14": ["B04_14", "B17", "NGHIEN_CUU_VIEN_14", "{{NGHIEN_CUU_VIEN_14}}"],
    "B04_15": ["B04_15", "B18", "NGHIEN_CUU_VIEN_15", "{{NGHIEN_CUU_VIEN_15}}"],
    "B04_16": ["B04_16", "B19", "NGHIEN_CUU_VIEN_16", "{{NGHIEN_CUU_VIEN_16}}"],
    "B04_17": ["B04_17", "B20", "NGHIEN_CUU_VIEN_17", "{{NGHIEN_CUU_VIEN_17}}"],
    "B04_18": ["B04_18", "NGHIEN_CUU_VIEN_18", "{{NGHIEN_CUU_VIEN_18}}"],
    "B04_19": ["B04_19", "NGHIEN_CUU_VIEN_19", "{{NGHIEN_CUU_VIEN_19}}"],
    "B04_20": ["B04_20", "NGHIEN_CUU_VIEN_20", "{{NGHIEN_CUU_VIEN_20}}"],
    "C01": ["C01", "CHU_TICH_HD_DAO_DUC", "{{CHU_TICH_HD_DAO_DUC}}"],
    "D01": ["D01", "CHU_TICH_HD_KHOA_HOC", "{{CHU_TICH_HD_KHOA_HOC}}"],
    "E01": ["E01", "CHU_TICH_HD_NGHIEM_THU", "{{CHU_TICH_HD_NGHIEM_THU}}", "CHU_TICH_HD_NGHIEM_THU_TEN", "{{CHU_TICH_HD_NGHIEM_THU_TEN}}"],
}


def _build_code_index(ws) -> dict:
    index = {}
    for row in ws.iter_rows(min_row=3, max_col=1):
        cell = row[0]
        code = cell.value
        if isinstance(code, str) and not code.startswith("SEC_"):
            raw_str = code.strip()
            index[raw_str] = cell.row
            clean_str = raw_str.replace("{{", "").replace("}}", "").strip()
            index[clean_str] = cell.row
            index[f"{{{{{clean_str}}}}}"] = cell.row
            if " - " in raw_str:
                parts = [p.strip() for p in raw_str.split(" - ")]
                for p in parts:
                    index[p] = cell.row
                    clean_p = p.replace("{{", "").replace("}}", "").strip()
                    index[clean_p] = cell.row
                    index[f"{{{{{clean_p}}}}}"] = cell.row
    return index


def _resolve_row(index: dict, code: str) -> Optional[int]:
    if code in index:
        return index[code]
    for k, aliases in CODE_ALIASES.items():
        if code == k or code in aliases:
            for alias in [k] + aliases:
                if alias in index:
                    return index[alias]
    clean = code.replace("{{", "").replace("}}", "").strip()
    if clean in index:
        return index[clean]
    if f"{{{{{clean}}}}}" in index:
        return index[f"{{{{{clean}}}}}"]
    return None


NHAN_SU_SHEET_NAME = "_NhanSu"


def _load_nhan_su_registry(wb) -> dict:
    if wb is None or NHAN_SU_SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[NHAN_SU_SHEET_NAME]
    registry = {}
    for row_i in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_i, column=1).value
        if raw_name is None or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()
        degree = ws.cell(row=row_i, column=2).value
        org = ws.cell(row=row_i, column=3).value
        registry[name] = (
            "" if degree is None else str(degree).strip(),
            "" if org is None else str(org).strip(),
        )
    return registry


def _cell_text(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()


def read_person(ws, index: dict, code: str, registry: Optional[dict] = None) -> Optional[Person]:
    row = _resolve_row(index, code)
    if row is None:
        raise KeyError(f"Không tìm thấy mã mục '{code}' trong checklist")
    name = _cell_text(ws, row, 3)
    if not name:
        return None
    degree = _cell_text(ws, row, 4)
    org = _cell_text(ws, row, 5)

    if registry is None and hasattr(ws, "parent") and ws.parent is not None:
        registry = _load_nhan_su_registry(ws.parent)

    if registry and name in registry:
        reg_degree, reg_org = registry[name]
        if not degree:
            degree = reg_degree
        if not org:
            org = reg_org

    return Person(name=name, degree=degree, org=org)


def read_text(ws, index: dict, code: str) -> str:
    row = _resolve_row(index, code)
    if row is None:
        raise KeyError(f"Không tìm thấy mã mục '{code}' trong checklist")
    return _cell_text(ws, row, 3)



def parse_committee(
    ws, index: dict, prefix: str, registry: Optional[dict] = None, required: bool = True
) -> CommitteeData:
    chair = read_person(ws, index, f"{prefix}01", registry=registry)
    if chair is None:
        if required:
            raise ValueError(f"Chủ tịch hội đồng ({prefix}01) là bắt buộc nhưng đang trống")
        chair = Person(name="", degree="", org="")

    reviewers = []
    for code in (f"{prefix}02", f"{prefix}03", f"{prefix}06"):
        person = read_person(ws, index, code, registry=registry)
        if person:
            reviewers.append(person)

    members = []
    for code in (f"{prefix}04", f"{prefix}05", f"{prefix}07", f"{prefix}08"):
        person = read_person(ws, index, code, registry=registry)
        if person:
            members.append(person)

    secretaries = []
    for code in (f"{prefix}09", f"{prefix}10"):
        person = read_person(ws, index, code, registry=registry)
        if person is None:
            if required:
                raise ValueError(f"Thư ký hội đồng ({code}) là bắt buộc nhưng đang trống")
            person = Person(name="", degree="", org="")
        secretaries.append(person)

    return CommitteeData(chair=chair, reviewers=reviewers, members=members, secretaries=secretaries)


def read_expert_cvs(ws, index: dict) -> List[ExpertCvEntry]:
    """Doc PHAN F, ma F02-F10 (F01 la chu nhiem, CV cua chu nhiem khop truc
    tiep qua info.head.name, khong doc rieng F01 nua). Dong nao ten (cot 3)
    de trong thi bo qua - khong con phu thuoc vao cot filename nua."""
    entries = []
    for i in range(2, 11):
        code = f"F{i:02d}"
        row = _resolve_row(index, code)
        if row is None:
            continue
        name = _cell_text(ws, row, 3)
        if not name:
            continue
        entries.append(ExpertCvEntry(code=code, name=name, role=_cell_text(ws, row, 4)))
    return entries


def read_researchers(ws, index: dict, registry: Optional[dict] = None) -> List[Person]:
    researchers = []
    for i in range(4, 21):
        code = f"B{i:02d}"
        row = _resolve_row(index, code)
        if row is None:
            continue
        person = read_person(ws, index, code, registry=registry)
        if person:
            researchers.append(person)
    return researchers


def load_project_data(xlsx_path: Path, sheet_name: str) -> ProjectInfo:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    index = _build_code_index(ws)
    registry = _load_nhan_su_registry(wb)

    # Import cuc bo de tranh vong lap import (token_rules import excel_reader
    # de dung read_text/read_person/parse_timeline).
    import token_rules
    common_tokens = token_rules.resolve_tokens(ws, index)

    title = read_text(ws, index, "A01")
    if not title:
        raise ValueError("Tên đề tài (A01) đang trống trong checklist")

    row_a03 = _resolve_row(index, "A03")
    if row_a03 is None:
        raise ValueError("Không tìm thấy dòng cho Năm thực hiện hồ sơ (A03)")
    year_raw = ws.cell(row=row_a03, column=3).value
    if year_raw is None:
        raise ValueError("Năm thực hiện hồ sơ (A03) đang trống trong checklist")

    head = read_person(ws, index, "B01", registry=registry)
    if head is None:
        raise ValueError("Chủ nhiệm đề tài (B01) là bắt buộc nhưng đang trống")

    researchers = read_researchers(ws, index, registry=registry)

    partner_org = read_text(ws, index, "A06") if "A06" in index else ""
    research_location = read_text(ws, index, "A07") if "A07" in index else ""

    pkg_row = _resolve_row(index, "CFG_PACKAGE") or _resolve_row(index, "BO_HO_SO") or _resolve_row(index, "A09")
    package_id = None
    if pkg_row:
        val = ws.cell(row=pkg_row, column=3).value
        package_id = str(val).strip() if val is not None and str(val).strip() else None

    # Xác định các section cần bắt buộc dựa trên package_id
    import dossier_packages
    active_sections = None
    if package_id:
        try:
            pkg = dossier_packages.get_package(package_id)
            active_sections = set(pkg.sections)
        except (KeyError, ValueError):
            pass

    ethics_required = active_sections is None or "dao_duc" in active_sections
    proposal_required = active_sections is None or "khoa_hoc" in active_sections
    acceptance_required = active_sections is None or "nghiem_thu" in active_sections

    ethics_committee = parse_committee(ws, index, "C", registry=registry, required=ethics_required)
    proposal_committee = parse_committee(ws, index, "D", registry=registry, required=proposal_required)
    acceptance_committee = parse_committee(ws, index, "E", registry=registry, required=acceptance_required)

    return ProjectInfo(
        title=title,
        research_type=read_text(ws, index, "A02"),
        year=int(year_raw),
        host_org=read_text(ws, index, "A04"),
        partner_org=partner_org or None,
        research_location=research_location or None,
        timeline=read_text(ws, index, "A05"),
        head=head,
        co_head=read_person(ws, index, "B02", registry=registry),
        project_secretary=read_person(ws, index, "B03", registry=registry),
        researchers=researchers,
        ethics_committee=ethics_committee,
        proposal_committee=proposal_committee,
        acceptance_committee=acceptance_committee,
        expert_cvs=read_expert_cvs(ws, index),
        package_id=package_id,
        common_tokens=common_tokens,
    )
