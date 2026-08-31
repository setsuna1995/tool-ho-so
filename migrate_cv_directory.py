# migrate_cv_directory.py
import sys
import shutil
from pathlib import Path
import openpyxl

if sys.stdout:
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

ROOT = Path(__file__).resolve().parent
OLD_DIR = ROOT / "CV chuyên gia"
NEW_DIR = ROOT / "Lý lịch khoa học"
CHECKLIST_PATH = ROOT / "Form checklist hồ sơ dự án.xlsx"


def migrate_folder() -> None:
    if OLD_DIR.exists() and not NEW_DIR.exists():
        OLD_DIR.rename(NEW_DIR)
        print(f"Da doi ten thu muc '{OLD_DIR.name}' -> '{NEW_DIR.name}'")
    elif not NEW_DIR.exists():
        NEW_DIR.mkdir(parents=True, exist_ok=True)
        print(f"Da tao thu muc '{NEW_DIR.name}'")

    # Xoa cac file TM-*.pdf mau cu khong phai CV
    for pdf_file in NEW_DIR.glob("TM-*.pdf"):
        pdf_file.unlink()
        print(f"Da xoa file thu moi PDF cu: {pdf_file.name}")
    for temp_file in NEW_DIR.glob("~$*"):
        temp_file.unlink()


def migrate_checklist() -> None:
    if not CHECKLIST_PATH.exists():
        return
    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    for sheet_name in ("Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for r in range(1, ws.max_row + 1):
                c1 = ws.cell(r, 1).value
                if c1 in ("F02", "F03", "F04", "F05", "F06", "F07", "F08", "F09", "F10"):
                    ws.cell(row=r, column=3).value = None
                    ws.cell(row=r, column=4).value = None
                    ws.cell(row=r, column=5).value = None
    wb.save(CHECKLIST_PATH)
    print("Da don dep cac dong F02-F10 trong checklist.")


def migrate_all() -> None:
    migrate_folder()
    migrate_checklist()


if __name__ == "__main__":
    migrate_all()
    print("XONG: Da to chuc lai thu muc Ly lich khoa hoc.")
