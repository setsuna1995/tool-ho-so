# migrate_clear_f03_missing_cv.py
from pathlib import Path

import openpyxl

CHECKLIST_PATH = Path(__file__).resolve().parent / "Form checklist hồ sơ dự án.xlsx"
SHEET_VIAM = "Đề tài - Bánh ăn dặm VIAM 2027"


def _find_row(ws, code: str) -> int:
    for row in ws.iter_rows(min_row=5, max_col=1):
        if row[0].value == code:
            return row[0].row
    raise ValueError(f"Không tìm thấy mã mục '{code}' trong sheet '{ws.title}'")


def clear_f03_pending_cv() -> None:
    """F03 (Luu Lien Huong, Thu ky De tai) da khai ten nhung chua co file CV
    khop trong 'CV chuyen gia/' - xoa ten/vai tro de tranh loi
    FileNotFoundError khi sinh ho so (spec §7). De khai lai: dat file CV
    that cua co vao 'CV chuyen gia/', roi go lai ten + vai tro vao F03
    trong sheet nay."""
    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    ws = wb[SHEET_VIAM]
    row = _find_row(ws, "F03")
    if ws.cell(row=row, column=3).value is None:
        return
    ws.cell(row=row, column=3).value = None
    ws.cell(row=row, column=4).value = None
    wb.save(CHECKLIST_PATH)


if __name__ == "__main__":
    clear_f03_pending_cv()
    print("Da xoa ten F03 (chua co file CV khop) khoi sheet du an VIAM 2027.")
