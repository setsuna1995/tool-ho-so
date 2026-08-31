# migrate_remove_cv_filename_column.py
from pathlib import Path

import openpyxl

CHECKLIST_PATH = Path(__file__).resolve().parent / "Form checklist hồ sơ dự án.xlsx"
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]
LISTS_SHEET = "_Lists"


def _build_code_index(ws) -> dict:
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str)
    }


def _clear_e_column_validations(ws, target_refs: set) -> None:
    keep = []
    for dv in ws.data_validations.dataValidation:
        dv_cells = set(str(dv.sqref).split())
        if not (dv_cells & target_refs):
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def _rewrite_status_formulas(ws, index: dict) -> None:
    f01_row = index.get("F01")
    if f01_row is not None:
        ws.cell(row=f01_row, column=6).value = (
            f'=IF(ISBLANK(C{f01_row}), "❌ CHƯA ĐIỀN THÔNG TIN CNĐT (BÁO LỖI)", '
            f'"✅ Đã khai chủ nhiệm đề tài - CV sẽ tự khớp theo tên")'
        )
    for i in range(2, 11):
        row = index.get(f"F{i:02d}")
        if row is None:
            continue
        ws.cell(row=row, column=6).value = (
            f'=IF(ISBLANK(C{row}), "⚪ Tùy chọn (Trống)", "✅ Đã khai tên - CV sẽ tự khớp theo tên")'
        )


def remove_cv_filename_column() -> None:
    wb = openpyxl.load_workbook(CHECKLIST_PATH)

    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        index = _build_code_index(ws)
        e_refs = {f"E{index[f'F{i:02d}']}" for i in range(1, 11) if f"F{i:02d}" in index}
        _clear_e_column_validations(ws, e_refs)
        _rewrite_status_formulas(ws, index)

    if LISTS_SHEET in wb.sheetnames:
        del wb[LISTS_SHEET]

    wb.save(CHECKLIST_PATH)


if __name__ == "__main__":
    remove_cv_filename_column()
    print("Da go cot 'TEN FILE CV' (dropdown + cong thuc trang thai) khoi ca 2 sheet.")
