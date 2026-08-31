import sys
from pathlib import Path

_SRC_DIR = Path(__file__).resolve().parent.parent / "src"
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

import excel_reader
import migrate_add_nhan_su_sheet as nhan_su

CHECKLIST_PATH = Path(__file__).resolve().parent.parent / excel_reader.CHECKLIST_FILENAME
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]
NHAN_SU_SHEET_NAME = nhan_su.NHAN_SU_SHEET_NAME
PERSON_CODE_PREFIXES = ("B", "C", "D", "E")


def _build_code_index(ws) -> dict:
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str) and not row[0].value.startswith("SEC_")
    }


def _nhan_su_row_count(wb) -> int:
    ws = wb[NHAN_SU_SHEET_NAME]
    return max(ws.max_row - 1, 0)


def _load_nhan_su_registry(wb) -> dict:
    """Doc sheet _NhanSu thanh {ten: (hoc_ham_hoc_vi, don_vi)}.

    Cot D/E cua sheet du an duoc dien bang GIA TRI that (Python tra cuu san),
    khong phai cong thuc VLOOKUP: excel_reader.load_project_data mo workbook
    voi data_only=True nen openpyxl khong bao gio tinh cong thuc - cong thuc
    se doc ra rong va lam mat hoc ham/don vi trong moi tai lieu sinh ra.
    """
    ws = wb[NHAN_SU_SHEET_NAME]
    registry = {}
    for row_i in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_i, column=1).value
        if raw_name is None or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()
        degree = ws.cell(row=row_i, column=2).value
        org = ws.cell(row=row_i, column=3).value
        registry[name] = (
            "" if degree is None else str(degree).strip(),
            "" if org is None else str(org).strip(),
        )
    return registry


def _clear_existing_person_validations(ws, target_refs: set) -> None:
    keep = []
    for dv in ws.data_validations.dataValidation:
        dv_cells = set(str(dv.sqref).split())
        if not (dv_cells & target_refs):
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def wire_person_dropdowns(checklist_path: Path = CHECKLIST_PATH) -> None:
    nhan_su.add_nhan_su_sheet(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    n = _nhan_su_row_count(wb)
    if n == 0:
        wb.save(checklist_path)
        return

    registry = _load_nhan_su_registry(wb)

    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        index = _build_code_index(ws)
        person_rows = [row for code, row in index.items() if code[0] in PERSON_CODE_PREFIXES]

        name_refs = {f"C{row}" for row in person_rows}
        _clear_existing_person_validations(ws, name_refs)

        name_dv = DataValidation(
            type="list", formula1=f"='{NHAN_SU_SHEET_NAME}'!$A$2:$A${n + 1}", allow_blank=True
        )
        ws.add_data_validation(name_dv)
        for row in person_rows:
            name_dv.add(ws.cell(row=row, column=3))
            raw_name = ws.cell(row=row, column=3).value
            name = "" if raw_name is None else str(raw_name).strip()
            if not name:
                ws.cell(row=row, column=4, value="")
                ws.cell(row=row, column=5, value="")
            else:
                degree, org = registry.get(name, ("", ""))
                if not ws.cell(row=row, column=4).value:
                    ws.cell(row=row, column=4, value=degree)
                if not ws.cell(row=row, column=5).value:
                    ws.cell(row=row, column=5, value=org)

    wb.save(checklist_path)


DEGREE_CHOICES = ["GS.TS.", "PGS.TS.", "TS.", "ThS.", "BS.CKII", "BS.CKI", "BS.", "CN.", "DS."]
RESEARCH_TYPE_CHOICES = [
    "Nghiên cứu can thiệp",
    "Nghiên cứu quan sát",
    "Thử nghiệm lâm sàng",
    "Đánh giá hiệu quả công thức",
    "Khác",
]
DOSSIER_PACKAGE_CHOICES = [
    "A",
    "B",
    "C",
    "D",
    "A_dao_duc",
    "A_khoa_hoc",
    "A_nghiem_thu",
]


TOKENS_SHEET_NAME = "_Tokens"

DEFAULT_TOKEN_LIST = [
    ("{{TEN_DE_TAI}}", "A01", "Tên đề tài nghiên cứu"),
    ("{{NAM}}", "A03", "Năm thực hiện hồ sơ"),
    ("{{DON_VI_CHU_TRI}}", "A04", "Cơ quan chủ trì"),
    ("{{DON_VI_DOI_TAC}}", "A06", "Cơ quan phối hợp (tùy chọn)"),
    ("{{THOI_GIAN_BAT_DAU}}", "A05", "Mốc thời gian bắt đầu (MM/YYYY)"),
    ("{{THOI_GIAN_KET_THUC}}", "A05", "Mốc thời gian kết thúc (MM/YYYY)"),
    ("{{DIA_DIEM_TRIEN_KHAI}}", "A07", "Địa điểm triển khai nghiên cứu"),
    ("{{DAU_MOI_LIEN_HE}}", "A08", "Đầu mối liên hệ thư mời"),
    ("{{CHU_NHIEM_HO_TEN}}", "B01", "Họ tên Chủ nhiệm đề tài"),
    ("{{CHU_NHIEM_TEN}}", "B01", "Tên riêng Chủ nhiệm đề tài"),
    ("{{CHU_NHIEM_DON_VI}}", "B01", "Đơn vị công tác Chủ nhiệm đề tài"),
    ("{{DONG_CHU_NHIEM_HO_TEN}}", "B02", "Họ tên Đồng chủ nhiệm đề tài"),
    ("{{THU_KY_DE_TAI}}", "B03", "Họ tên Thư ký đề tài"),
    ("{{DANH_SACH_NGHIEN_CUU_VIEN}}", "B04-B20", "Danh sách nghiên cứu viên (có đánh số)"),
    ("{{NGHIEN_CUU_VIEN_1}}", "B04", "Nghiên cứu viên 1 - Họ tên có học vị"),
    ("{{NGHIEN_CUU_VIEN_1_TEN}}", "B04", "Nghiên cứu viên 1 - Chỉ họ tên"),
    ("{{NGHIEN_CUU_VIEN_1_DON_VI}}", "B04", "Nghiên cứu viên 1 - Đơn vị công tác"),
    ("{{NGHIEN_CUU_VIEN_2}}", "B05", "Nghiên cứu viên 2 - Họ tên có học vị"),
    ("{{NGHIEN_CUU_VIEN_2_TEN}}", "B05", "Nghiên cứu viên 2 - Chỉ họ tên"),
    ("{{NGHIEN_CUU_VIEN_2_DON_VI}}", "B05", "Nghiên cứu viên 2 - Đơn vị công tác"),
    ("{{NGHIEN_CUU_VIEN_3}}", "B06", "Nghiên cứu viên 3 - Họ tên có học vị"),
    ("{{NGHIEN_CUU_VIEN_3_TEN}}", "B06", "Nghiên cứu viên 3 - Chỉ họ tên"),
    ("{{NGHIEN_CUU_VIEN_3_DON_VI}}", "B06", "Nghiên cứu viên 3 - Đơn vị công tác"),
    ("{{NGHIEN_CUU_VIEN_4}}", "B07", "Nghiên cứu viên 4 - Họ tên có học vị"),
    ("{{NGHIEN_CUU_VIEN_5}}", "B08", "Nghiên cứu viên 5 - Họ tên có học vị"),
    ("{{CHU_TICH_HD_DAO_DUC}}", "C01", "Chủ tịch HĐ Đạo đức"),
    ("{{CHU_TICH_HD_KHOA_HOC}}", "D01", "Chủ tịch HĐ Khoa học"),
    ("{{CHU_TICH_HD_NGHIEM_THU}}", "E01", "Chủ tịch HĐ Nghiệm thu"),
    ("{{BOI_CANH_DU_AN}}", "G01", "Bối cảnh / lý do triển khai đề tài"),
    ("{{DIA_DIEM_HOP}}", "G02", "Địa điểm họp Hội đồng"),
    ("{{THOI_GIAN_HOP}}", "G03", "Thời gian họp Hội đồng (giờ, ngày tháng)"),
    ("{{SO_CONG_VAN}}", "G04", "Số công văn mời chuyên gia"),
    ("{{NGAY_CONG_VAN}}", "G05", "Ngày ký công văn mời chuyên gia"),
    ("{{SO_QD_GIAO_DE_TAI}}", "G06", "Số Quyết định giao đề tài (00)"),
    ("{{NGAY_QD_GIAO_DE_TAI}}", "G07", "Ngày ký QĐ giao đề tài"),
    ("{{SO_QD_TLHD_DAO_DUC}}", "G08", "Số QĐ TLHĐ Đạo đức (01)"),
    ("{{NGAY_QD_TLHD_DAO_DUC}}", "G09", "Ngày ký QĐ TLHĐ Đạo đức"),
    ("{{SO_QD_TLHD_KHOA_HOC}}", "G10", "Số QĐ TLHĐ Khoa học (05)"),
    ("{{NGAY_QD_TLHD_KHOA_HOC}}", "G11", "Ngày ký QĐ TLHĐ Khoa học"),
    ("{{SO_QD_TLHD_PHE_DUYET}}", "G14", "Số QĐ Phê duyệt đề tài (08)"),
    ("{{NGAY_QD_PHE_DUYET}}", "G15", "Ngày ký QĐ Phê duyệt đề tài"),
    ("{{SO_QD_TLHD_NGHIEM_THU}}", "G12", "Số QĐ TLHĐ Nghiệm thu (9)"),
    ("{{NGAY_QD_TLHD_NGHIEM_THU}}", "G13", "Ngày ký QĐ TLHĐ Nghiệm thu"),
    ("{{SO_QD_CONG_NHAN}}", "G16", "Số QĐ Công nhận kết quả (12)"),
    ("{{NGAY_QD_CONG_NHAN}}", "G17", "Ngày ký QĐ Công nhận kết quả"),
    ("{{NGAY_HOP_DAO_DUC}}", "G18", "Ngày họp HĐ Đạo đức (02, 03)"),
    ("{{NGAY_HOP_KHOA_HOC}}", "G19", "Ngày họp HĐ Khoa học (06, 07)"),
    ("{{NGAY_HOP_NGHIEM_THU}}", "G20", "Ngày họp HĐ Nghiệm thu (10, 11)"),
]


def sync_tokens_sheet(wb) -> None:
    if TOKENS_SHEET_NAME in wb.sheetnames:
        ws = wb[TOKENS_SHEET_NAME]
    else:
        ws = wb.create_sheet(TOKENS_SHEET_NAME)

    ws.cell(row=1, column=1, value="TOKEN_TAG")
    ws.cell(row=1, column=2, value="MÃ_MỤC")
    ws.cell(row=1, column=3, value="MÔ_TẢ")

    for i, (tok, code, desc) in enumerate(DEFAULT_TOKEN_LIST, start=2):
        ws.cell(row=i, column=1, value=tok)
        ws.cell(row=i, column=2, value=code)
        ws.cell(row=i, column=3, value=desc)


def wire_token_dropdowns(checklist_path: Path = CHECKLIST_PATH) -> None:
    wb = openpyxl.load_workbook(checklist_path)
    sync_tokens_sheet(wb)
    num_tokens = len(DEFAULT_TOKEN_LIST)

    target_sheets = [s for s in SHEET_NAMES if s in wb.sheetnames]
    if not target_sheets:
        target_sheets = [s for s in wb.sheetnames if s not in (NHAN_SU_SHEET_NAME, TOKENS_SHEET_NAME)]

    token_formula = f"='{TOKENS_SHEET_NAME}'!$A$2:$A${num_tokens + 1}"

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        col_a_refs = {f"A{r}" for r in range(6, 60)}
        _clear_existing_person_validations(ws, col_a_refs)

        tok_dv = DataValidation(type="list", formula1=token_formula, allow_blank=True)
        ws.add_data_validation(tok_dv)
        for r in range(6, 60):
            val_a = ws.cell(row=r, column=1).value
            if val_a and not str(val_a).startswith("SEC_"):
                tok_dv.add(ws.cell(row=r, column=1))

    wb.save(checklist_path)


BO_HO_SO_SHEET_NAME = "_BoHoSo"


def sync_bo_ho_so_sheet(wb) -> int:
    """Tạo và đồng bộ sheet _BoHoSo với danh mục các Bộ hồ sơ trong hệ thống."""
    import dossier_packages
    packages = dossier_packages.get_available_packages()

    if BO_HO_SO_SHEET_NAME in wb.sheetnames:
        ws = wb[BO_HO_SO_SHEET_NAME]
    else:
        ws = wb.create_sheet(BO_HO_SO_SHEET_NAME)

    ws.cell(row=1, column=1, value="MÃ_BỘ")
    ws.cell(row=1, column=2, value="TÊN_BỘ_HỒ_SƠ")
    ws.cell(row=1, column=3, value="MÔ_TẢ_PHẠM_VI")
    ws.cell(row=1, column=4, value="CÁC_PHẦN_XUẤT")
    ws.cell(row=1, column=5, value="THƯ_MỤC_MẪU")

    for i, pkg in enumerate(packages, start=2):
        ws.cell(row=i, column=1, value=pkg.id)
        ws.cell(row=i, column=2, value=pkg.name)
        ws.cell(row=i, column=3, value=pkg.description)
        ws.cell(row=i, column=4, value=", ".join(pkg.sections))
        ws.cell(row=i, column=5, value=", ".join(pkg.template_dirs))

    return len(packages)


def wire_extended_validations(checklist_path: Path = CHECKLIST_PATH, protect: bool = False) -> None:
    from openpyxl.styles import Protection

    wire_person_dropdowns(checklist_path)
    wire_token_dropdowns(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    num_pkgs = sync_bo_ho_so_sheet(wb)

    target_sheets = [s for s in SHEET_NAMES if s in wb.sheetnames]
    if not target_sheets:
        target_sheets = [s for s in wb.sheetnames if s not in (NHAN_SU_SHEET_NAME, TOKENS_SHEET_NAME, BO_HO_SO_SHEET_NAME)]

    degree_formula = f'"{",".join(DEGREE_CHOICES)}"'
    research_formula = f'"{",".join(RESEARCH_TYPE_CHOICES)}"'
    pkg_formula = f"='{BO_HO_SO_SHEET_NAME}'!$A$2:$A${num_pkgs + 1}"

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        index = _build_code_index(ws)

        # 0. Dropdown cho CFG_PACKAGE / Header Row 3
        pkg_row = index.get("CFG_PACKAGE") or index.get("BO_HO_SO") or 3
        _clear_existing_person_validations(ws, {f"C{pkg_row}"})
        pkg_dv = DataValidation(type="list", formula1=pkg_formula, allow_blank=True)
        ws.add_data_validation(pkg_dv)
        pkg_dv.add(ws.cell(row=pkg_row, column=3))

        # 1. Dropdown cho A02 (Loại hình nghiên cứu)
        if "A02" in index:
            a02_row = index["A02"]
            _clear_existing_person_validations(ws, {f"C{a02_row}"})
            rt_dv = DataValidation(type="list", formula1=research_formula, allow_blank=True)
            ws.add_data_validation(rt_dv)
            rt_dv.add(ws.cell(row=a02_row, column=3))

        # 2. Dropdown cho cột Degree (D) của các dòng nhân sự
        person_rows = [row for code, row in index.items() if code[0] in PERSON_CODE_PREFIXES]
        degree_refs = {f"D{row}" for row in person_rows}
        _clear_existing_person_validations(ws, degree_refs)
        deg_dv = DataValidation(type="list", formula1=degree_formula, allow_blank=True)
        ws.add_data_validation(deg_dv)
        for row in person_rows:
            deg_dv.add(ws.cell(row=row, column=4))

        # 3. Mở khóa ô nhập liệu & Protect Sheet nếu được yêu cầu
        if protect:
            unlocked = Protection(locked=False)
            for row in index.values():
                for col in (3, 4, 5):
                    ws.cell(row=row, column=col).protection = unlocked
            ws.protection.sheet = True

    wb.save(checklist_path)


if __name__ == "__main__":
    wire_extended_validations(protect=False)
    print("Da gan dropdown token, ten, hoc vi, loai hinh nghien cuu cho toan bo checklist.")


