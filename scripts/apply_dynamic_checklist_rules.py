# scripts/apply_dynamic_checklist_rules.py
import sys
from pathlib import Path

_SRC_DIR = Path(__file__).resolve().parent.parent / "src"
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

import openpyxl
import excel_reader
import capnhat_nhan_su

CHECKLIST_PATH = Path(__file__).resolve().parent.parent / excel_reader.CHECKLIST_FILENAME

ETHICS_REQUIRED = {"C01", "C02", "C03", "C04", "C05", "C09", "C10", "CHU_TICH_HD_DAO_DUC"}
PROPOSAL_REQUIRED = {"D01", "D02", "D03", "D04", "D05", "D09", "D10", "CHU_TICH_HD_KHOA_HOC"}
ACCEPTANCE_REQUIRED = {"E01", "E02", "E03", "E04", "E05", "E09", "E10", "CHU_TICH_HD_NGHIEM_THU"}
GENERAL_REQUIRED = {
    "A01", "A02", "A03", "A04", "A05", "B01", "F01",
    "TEN_DE_TAI", "LOAI_HINH_NC", "NAM", "DON_VI_CHU_TRI", "THOI_GIAN_BAT_DAU", "CHU_NHIEM_HO_TEN"
}


def apply_dynamic_rules(checklist_path: Path = CHECKLIST_PATH) -> None:
    wb = openpyxl.load_workbook(checklist_path)
    capnhat_nhan_su.sync_bo_ho_so_sheet(wb)

    for s in ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]

        # 1. Cấu hình Header Row 3
        ws.cell(row=3, column=1, value="BO_HO_SO")
        ws.cell(row=3, column=2, value="CHỌN BỘ HỒ SƠ (A: NCKH & Đánh giá công thức / B: Thử nghiệm lâm sàng / C: Quan sát dịch tễ / D: Chuyển giao CN)")
        if not ws.cell(row=3, column=3).value:
            ws.cell(row=3, column=3, value="A")

        # 2. Banner kiểm tra tổng thể Row 2
        ws.cell(
            row=2,
            column=1,
            value='=IF(COUNTIF(F6:F150, "*BÁO LỖI*")>0, "⚠️ BÁO LỖI: PHÁT HIỆN THIẾU THÔNG TIN BẮT BUỘC CỦA BỘ HỒ SƠ ĐÃ CHỌN - VUI LÒNG BỔ SUNG CÁC Ô MÀU ĐỎ DƯỚI ĐÂY", "🎉 HOÀN THÀNH 100%: TẤT CẢ THÔNG TIN BẮT BUỘC CHO BỘ HỒ SƠ ĐÃ CHỌN ĐÃ ĐẦY ĐỦ - SẴN SÀNG CHẠY RA BỘ HỒ SƠ")',
        )

        # 3. Quét từng dòng từ Row 5
        for r in range(5, ws.max_row + 1):
            cell_a = ws.cell(row=r, column=1).value
            if not cell_a:
                continue
            raw_a = str(cell_a).strip()
            clean_a = raw_a.replace("{{", "").replace("}}", "").strip()

            # Section Headers
            if raw_a == "SEC_A":
                ws.cell(row=r, column=2, value="PHẦN A: THÔNG TIN CHUNG DỰ ÁN & ĐỀ TÀI (BẮT BUỘC TẤT CẢ CÁC BỘ)")
            elif raw_a == "SEC_B":
                ws.cell(row=r, column=2, value="PHẦN B: BAN CHỦ NHIỆM & NGHIÊN CỨU VIÊN ĐỀ TÀI (CHỦ NHIỆM BẮT BUỘC)")
            elif raw_a == "SEC_C":
                ws.cell(
                    row=r,
                    column=2,
                    value='="PHẦN C: HỘI ĐỒNG ĐẠO ĐỨC TRONG NGHIÊN CỨU " & IF(ISNUMBER(SEARCH("dao_duc", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), "🟢 [BẮT BUỘC - ĐANG ÁP DỤNG]", "⚪ [KHÔNG ÁP DỤNG CHO BỘ NÀY]")',
                )
            elif raw_a == "SEC_D":
                ws.cell(
                    row=r,
                    column=2,
                    value='="PHẦN D: HỘI ĐỒNG KHOA HỌC XÉT DUYỆT ĐỀ CƯƠNG " & IF(ISNUMBER(SEARCH("khoa_hoc", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), "🟢 [BẮT BUỘC - ĐANG ÁP DỤNG]", "⚪ [KHÔNG ÁP DỤNG CHO BỘ NÀY]")',
                )
            elif raw_a == "SEC_E":
                ws.cell(
                    row=r,
                    column=2,
                    value='="PHẦN E: HỘI ĐỒNG ĐÁNH GIÁ NGHIỆM THU " & IF(ISNUMBER(SEARCH("nghiem_thu", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), "🟢 [BẮT BUỘC - ĐANG ÁP DỤNG]", "⚪ [KHÔNG ÁP DỤNG CHO BỘ NÀY]")',
                )
            elif raw_a == "SEC_F":
                ws.cell(
                    row=r,
                    column=2,
                    value='="PHẦN F: LÝ LỊCH KHOA HỌC CHUYÊN GIA (CV) " & IF(ISNUMBER(SEARCH("dao_duc", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), "🟢 [BẮT BUỘC CV CHỦ NHIỆM]", "⚪ [TÙY CHỌN]")',
                )
            elif raw_a.startswith("SEC_"):
                continue
            else:
                # Dòng item: Gán công thức Cột F kiểm tra thông minh
                if clean_a in GENERAL_REQUIRED or raw_a in GENERAL_REQUIRED:
                    ws.cell(row=r, column=6, value=f'=IF(ISBLANK(C{r}), "❌ CHƯA ĐIỀN (BÁO LỖI)", "✅ Xong")')
                elif clean_a in ETHICS_REQUIRED or raw_a in ETHICS_REQUIRED:
                    ws.cell(
                        row=r,
                        column=6,
                        value=f'=IF(ISNUMBER(SEARCH("dao_duc", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), IF(ISBLANK(C{r}), "❌ CHƯA ĐIỀN (BÁO LỖI)", "✅ Xong"), "⚪ Không yêu cầu cho bộ này")',
                    )
                elif clean_a in PROPOSAL_REQUIRED or raw_a in PROPOSAL_REQUIRED:
                    ws.cell(
                        row=r,
                        column=6,
                        value=f'=IF(ISNUMBER(SEARCH("khoa_hoc", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), IF(ISBLANK(C{r}), "❌ CHƯA ĐIỀN (BÁO LỖI)", "✅ Xong"), "⚪ Không yêu cầu cho bộ này")',
                    )
                elif clean_a in ACCEPTANCE_REQUIRED or raw_a in ACCEPTANCE_REQUIRED:
                    ws.cell(
                        row=r,
                        column=6,
                        value=f'=IF(ISNUMBER(SEARCH("nghiem_thu", VLOOKUP($C$3, _BoHoSo!$A$2:$D$8, 4, FALSE))), IF(ISBLANK(C{r}), "❌ CHƯA ĐIỀN (BÁO LỖI)", "✅ Xong"), "⚪ Không yêu cầu cho bộ này")',
                    )
                else:
                    ws.cell(row=r, column=6, value=f'=IF(ISBLANK(C{r}), "⚪ Tùy chọn (Trống)", "✅ Xong")')

    wb.save(checklist_path)
    # Gắn lại toàn bộ Dropdown
    capnhat_nhan_su.wire_extended_validations(checklist_path)


if __name__ == "__main__":
    apply_dynamic_rules()
    print("Da cap nhat toan bo cong thuc va giao dien thich ung dong theo Bo ho so thanh cong!")
