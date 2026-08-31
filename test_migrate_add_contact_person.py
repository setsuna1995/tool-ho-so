# test_migrate_add_contact_person.py
from pathlib import Path

import openpyxl
import pytest

import migrate_add_contact_person as migrate
import excel_reader

CHECKLIST_PATH = Path(__file__).resolve().parent / excel_reader.CHECKLIST_FILENAME
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]


def test_add_contact_person_field_adds_a08_to_both_sheets():
    migrate.add_contact_person_field()

    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        codes = [row[0].value for row in ws.iter_rows(min_row=5, max_col=1)]
        assert "A08" in codes


def test_add_contact_person_field_is_idempotent():
    migrate.add_contact_person_field()
    migrate.add_contact_person_field()

    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        codes = [row[0].value for row in ws.iter_rows(min_row=5, max_col=1)]
        assert codes.count("A08") == 1


def _code_index(ws):
    index = {}
    for row in ws.iter_rows(min_row=5, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and not cell.value.startswith("SEC_"):
            index[cell.value] = cell.row
    return index


@pytest.mark.parametrize("sheet_name", SHEET_NAMES)
def test_shifted_rows_reference_their_own_row(sheet_name):
    migrate.add_contact_person_field()

    wb = openpyxl.load_workbook(CHECKLIST_PATH, data_only=False)
    ws = wb[sheet_name]
    index = _code_index(ws)
    b01_row = index["B01"]
    formula = ws.cell(row=b01_row, column=6).value
    assert f"C{b01_row}" in formula
