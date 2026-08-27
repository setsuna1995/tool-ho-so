from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

import excel_reader

CHECKLIST_PATH = Path(__file__).resolve().parent / excel_reader.CHECKLIST_FILENAME
CV_DIR = Path(__file__).resolve().parent / "CV chuyên gia"
TEMPLATE_SHEET = "Đề tài - Mẫu trắng dự án mới"
LISTS_SHEET = "_Lists"
RESEARCH_TYPES = ["TVCT_ĐGHQ", "TNLS"]


def _refresh_lists_sheet(wb, cv_filenames: list) -> int:
    if LISTS_SHEET in wb.sheetnames:
        del wb[LISTS_SHEET]
    ws = wb.create_sheet(LISTS_SHEET)
    ws.sheet_state = "hidden"
    for i, name in enumerate(sorted(cv_filenames), start=1):
        ws.cell(row=i, column=1, value=name)
    return len(cv_filenames)


def _clear_existing_validations(ws, cell_refs: set) -> None:
    keep = []
    for dv in ws.data_validations.dataValidation:
        if not any(ref in str(dv.sqref) for ref in cell_refs):
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def _build_code_index(ws) -> dict:
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str)
    }


def refresh(checklist_path: Path = CHECKLIST_PATH, cv_dir: Path = CV_DIR) -> None:
    cv_filenames = [p.name for p in cv_dir.iterdir() if p.is_file()]

    wb = openpyxl.load_workbook(checklist_path)
    n = _refresh_lists_sheet(wb, cv_filenames)

    ws = wb[TEMPLATE_SHEET]
    index = _build_code_index(ws)

    f_cell_refs = {f"E{index[f'F{i:02d}']}" for i in range(1, 11) if f"F{i:02d}" in index}
    a02_cell_ref = {f"C{index['A02']}"} if "A02" in index else set()
    _clear_existing_validations(ws, f_cell_refs | a02_cell_ref)

    if n > 0:
        cv_dv = DataValidation(type="list", formula1=f"='{LISTS_SHEET}'!$A$1:$A${n}", allow_blank=True)
        ws.add_data_validation(cv_dv)
        for ref in f_cell_refs:
            cv_dv.add(ws[ref])

    if a02_cell_ref:
        research_type_dv = DataValidation(
            type="list", formula1=f'"{",".join(RESEARCH_TYPES)}"', allow_blank=True
        )
        ws.add_data_validation(research_type_dv)
        for ref in a02_cell_ref:
            research_type_dv.add(ws[ref])

    wb.save(checklist_path)


if __name__ == "__main__":
    refresh()
    print("Da cap nhat danh sach dropdown CV va kieu nghien cuu trong checklist.")
