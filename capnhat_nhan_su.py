from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

import excel_reader
import migrate_add_nhan_su_sheet as nhan_su

CHECKLIST_PATH = Path(__file__).resolve().parent / excel_reader.CHECKLIST_FILENAME
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]
NHAN_SU_SHEET_NAME = nhan_su.NHAN_SU_SHEET_NAME
PERSON_CODE_PREFIXES = ("B", "C", "D", "E")


def _build_code_index(ws) -> dict:
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str) and not row[0].value.startswith("SEC_")
    }


def _nhan_su_row_count(wb) -> int:
    ws = wb[NHAN_SU_SHEET_NAME]
    return max(ws.max_row - 1, 0)


def _load_nhan_su_registry(wb) -> dict:
    """Doc sheet _NhanSu thanh {ten: (hoc_ham_hoc_vi, don_vi)}.

    Cot D/E cua sheet du an duoc dien bang GIA TRI that (Python tra cuu san),
    khong phai cong thuc VLOOKUP: excel_reader.load_project_data mo workbook
    voi data_only=True nen openpyxl khong bao gio tinh cong thuc - cong thuc
    se doc ra rong va lam mat hoc ham/don vi trong moi tai lieu sinh ra.
    """
    ws = wb[NHAN_SU_SHEET_NAME]
    registry = {}
    for row_i in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_i, column=1).value
        if raw_name is None or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()
        degree = ws.cell(row=row_i, column=2).value
        org = ws.cell(row=row_i, column=3).value
        registry[name] = (
            "" if degree is None else str(degree).strip(),
            "" if org is None else str(org).strip(),
        )
    return registry


def _clear_existing_person_validations(ws, target_refs: set) -> None:
    keep = []
    for dv in ws.data_validations.dataValidation:
        dv_cells = set(str(dv.sqref).split())
        if not (dv_cells & target_refs):
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def wire_person_dropdowns(checklist_path: Path = CHECKLIST_PATH) -> None:
    nhan_su.add_nhan_su_sheet(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    n = _nhan_su_row_count(wb)
    if n == 0:
        wb.save(checklist_path)
        return

    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        index = _build_code_index(ws)
        person_rows = [row for code, row in index.items() if code[0] in PERSON_CODE_PREFIXES]

        name_refs = {f"C{row}" for row in person_rows}
        _clear_existing_person_validations(ws, name_refs)

        name_dv = DataValidation(
            type="list", formula1=f"='{NHAN_SU_SHEET_NAME}'!$A$2:$A${n + 1}", allow_blank=True
        )
        ws.add_data_validation(name_dv)
        for row in person_rows:
            name_dv.add(ws.cell(row=row, column=3))
            ws.cell(
                row=row,
                column=4,
                value=f'=IFERROR(VLOOKUP(C{row},{NHAN_SU_SHEET_NAME}!$A:$C,2,FALSE),"")',
            )
            ws.cell(
                row=row,
                column=5,
                value=f'=IFERROR(VLOOKUP(C{row},{NHAN_SU_SHEET_NAME}!$A:$C,3,FALSE),"")',
            )

    wb.save(checklist_path)


if __name__ == "__main__":
    wire_person_dropdowns()
    print("Da gan dropdown chon ten + cong thuc tra cuu hoc ham/don vi cho toan bo dong nhan su.")
