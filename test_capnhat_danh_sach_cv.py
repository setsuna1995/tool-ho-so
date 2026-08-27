from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

import capnhat_danh_sach_cv as refresh


def _make_checklist(tmp_path, cv_filenames):
    cv_dir = tmp_path / "CV chuyên gia"
    cv_dir.mkdir()
    for name in cv_filenames:
        (cv_dir / name).write_text("x")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đề tài - Mẫu trắng dự án mới"
    ws.cell(row=6, column=1, value="A01")
    ws.cell(row=7, column=1, value="A02")
    ws.cell(row=67, column=1, value="F01")
    ws.cell(row=68, column=1, value="F02")
    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    return path, cv_dir


def test_refresh_writes_cv_filenames_into_hidden_lists_sheet(tmp_path):
    checklist_path, cv_dir = _make_checklist(tmp_path, ["a.docx", "b.pdf"])

    refresh.refresh(checklist_path, cv_dir)

    wb = openpyxl.load_workbook(checklist_path)
    assert "_Lists" in wb.sheetnames
    values = {c.value for c in wb["_Lists"]["A"] if c.value}
    assert values == {"a.docx", "b.pdf"}


def test_refresh_adds_data_validation_to_f01_and_f02_filename_cells(tmp_path):
    checklist_path, cv_dir = _make_checklist(tmp_path, ["a.docx"])

    refresh.refresh(checklist_path, cv_dir)

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]
    ranges = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert any("E67" in r for r in ranges)
    assert any("E68" in r for r in ranges)


def test_refresh_adds_static_data_validation_to_a02(tmp_path):
    checklist_path, cv_dir = _make_checklist(tmp_path, [])

    refresh.refresh(checklist_path, cv_dir)

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]
    a02_validations = [dv for dv in ws.data_validations.dataValidation if "C7" in str(dv.sqref)]
    assert len(a02_validations) == 1
    assert "TVCT_ĐGHQ" in a02_validations[0].formula1
    assert "TNLS" in a02_validations[0].formula1


def test_refresh_is_idempotent_on_lists_sheet(tmp_path):
    checklist_path, cv_dir = _make_checklist(tmp_path, ["a.docx"])

    refresh.refresh(checklist_path, cv_dir)
    refresh.refresh(checklist_path, cv_dir)

    wb = openpyxl.load_workbook(checklist_path)
    assert wb.sheetnames.count("_Lists") == 1


def test_refresh_does_not_delete_unrelated_validation_on_superstring_cell(tmp_path):
    """C70 is a valid, in-range cell that happens to contain "C7" (A02's own
    target ref) as a substring. A naive `ref in str(dv.sqref)` match would
    delete a hand-added validation on C70 every time someone reruns the
    refresh. It must survive untouched, while the script's own A02/F-cell
    validations still get correctly replaced (not duplicated) on rerun."""
    checklist_path, cv_dir = _make_checklist(tmp_path, ["a.docx"])

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]
    unrelated_dv = DataValidation(type="list", formula1='"X,Y"', allow_blank=True)
    ws.add_data_validation(unrelated_dv)
    unrelated_dv.add(ws["C70"])
    wb.save(checklist_path)

    refresh.refresh(checklist_path, cv_dir)
    refresh.refresh(checklist_path, cv_dir)

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]

    survivors = [dv for dv in ws.data_validations.dataValidation if "C70" in str(dv.sqref).split()]
    assert len(survivors) == 1
    assert survivors[0].formula1 == '"X,Y"'

    a02_validations = [dv for dv in ws.data_validations.dataValidation if "C7" in str(dv.sqref).split()]
    assert len(a02_validations) == 1
