import io
import shutil
import sys
import tempfile
from pathlib import Path

# Đảm bảo import được các module trong src/
_SRC_DIR = Path(__file__).resolve().parent / "src"
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

import openpyxl

import cv_matching
import dossier_packages
import excel_reader
import paths
import section_engine
import template_config
import tokens
import word_writer

PROJECT_SHEET_PREFIX = "Đề tài - "

ILLEGAL_FOLDER_CHARS = ':/\\*?"<>|'


def list_project_sheets(xlsx_path: Path) -> list:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        return [name for name in wb.sheetnames if name.startswith(PROJECT_SHEET_PREFIX)]
    finally:
        wb.close()


def choose_sheet_name(xlsx_path: Path) -> str:
    if len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        return sys.argv[1]

    sheets = list_project_sheets(xlsx_path)
    if not sheets:
        raise RuntimeError(
            f"Khong tim thay sheet nao bat dau bang '{PROJECT_SHEET_PREFIX}' trong file {xlsx_path.name}."
        )
    if len(sheets) == 1:
        print(f"Dung sheet du an: '{sheets[0]}'")
        return sheets[0]

    print("Chon sheet du an muon tao ho so:")
    for i, name in enumerate(sheets, start=1):
        print(f"  {i}. {name}")

    while True:
        raw = input(f"Nhap so thu tu (1-{len(sheets)}): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(sheets):
            return sheets[int(raw) - 1]
        print(f"  [LOI] Vui long nhap so tu 1 den {len(sheets)}.")


def choose_package() -> dossier_packages.DossierPackage:
    packages = dossier_packages.get_available_packages()
    print("\nChon bo ho so muon tao:")
    for i, pkg in enumerate(packages, start=1):
        default_tag = " [Mac dinh]" if pkg.id == "full" else ""
        print(f"  {i}. {pkg.name}{default_tag}")
        print(f"     -> {pkg.description}")

    while True:
        raw = input(f"Nhap so thu tu (1-{len(packages)}, nhan Enter de chon [1]): ").strip()
        if not raw:
            return dossier_packages.get_package("full")
        if raw.isdigit() and 1 <= int(raw) <= len(packages):
            return packages[int(raw) - 1]
        print(f"  [LOI] Vui long nhap so tu 1 den {len(packages)}.")


def parse_cli_args(xlsx_path: Path) -> tuple:
    """Tra ve (sheet_name, package) dua vao dong lenh hoac hoi truc tiep."""
    sheet_name = None
    package_id = None

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        arg = args[i]
        if arg in ("--package", "-p") and i + 1 < len(args):
            package_id = args[i + 1]
            i += 2
        elif not arg.startswith("-") and sheet_name is None:
            sheet_name = arg
            i += 1
        else:
            i += 1

    if sheet_name is None:
        sheet_name = choose_sheet_name(xlsx_path)

    if package_id:
        package = dossier_packages.get_package(package_id)
    else:
        package = choose_package()

    return sheet_name, package


def _get_allowed_template_dirs(package: dossier_packages.DossierPackage) -> set:
    """Lay danh sach template_dir duoc phep tu section_config, ket hop voi package.template_dirs."""
    allowed = set()
    if package.template_dirs:
        allowed.update(package.template_dirs)
    for section_id in package.sections:
        try:
            cfg = section_engine.load_section_config(section_id)
            allowed.add(cfg.template_dir)
        except KeyError:
            pass
    return allowed if allowed else None


def copy_templates(root: Path, dest_root: Path, package: dossier_packages.DossierPackage = None) -> None:
    copies = template_config.discover_copies(root)
    allowed_dirs = _get_allowed_template_dirs(package) if package else None

    for rel_src, rel_dst in copies:
        if allowed_dirs is not None:
            # Kiem tra xem file co thuoc thu muc mau duoc phep khong
            src_top = rel_src.replace("templates/", "").replace("templates\\", "").split("/")[0].split("\\")[0]
            if src_top not in allowed_dirs:
                continue

        src = root / rel_src
        dst = dest_root / rel_dst
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_bytes(src.read_bytes())


def copy_head_cv(root: Path, dest_root: Path, info) -> None:
    cv_directory = paths.cv_dir()
    src = cv_matching.find_cv_file(cv_directory, info.head.name, context=" (chủ nhiệm đề tài)")
    dst = dest_root / "01. Hồ sơ đạo đức đề cương" / src.name
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_bytes(src.read_bytes())


def copy_expert_cvs(root: Path, dest_root: Path, info, package: dossier_packages.DossierPackage = None) -> None:
    cv_directory = paths.cv_dir()
    resolved = [
        (entry, cv_matching.find_cv_file(cv_directory, entry.name, context=f" (mã mục {entry.code} - {entry.role})"))
        for entry in info.expert_cvs
    ]

    for entry, src in resolved:
        target_dir = None
        code_num = int(entry.code[1:]) if entry.code[1:].isdigit() else 2
        if code_num <= 5 and (dest_root / "01. Hồ sơ đạo đức đề cương").exists():
            target_dir = dest_root / "01. Hồ sơ đạo đức đề cương"
        elif code_num <= 8 and (dest_root / "02. Hồ sơ khoa học đề cương").exists():
            target_dir = dest_root / "02. Hồ sơ khoa học đề cương"
        elif (dest_root / "04. Hồ sơ nghiệm thu").exists():
            target_dir = dest_root / "04. Hồ sơ nghiệm thu"
        else:
            for sub in ("01. Hồ sơ đạo đức đề cương", "02. Hồ sơ khoa học đề cương", "04. Hồ sơ nghiệm thu"):
                if (dest_root / sub).exists():
                    target_dir = dest_root / sub
                    break

        if target_dir:
            dst = target_dir / src.name
            dst.parent.mkdir(parents=True, exist_ok=True)
            dst.write_bytes(src.read_bytes())


def _sanitize_folder_name(name: str) -> str:
    """Thay cac ky tu khong hop le trong ten thu muc Windows bang khoang trang."""
    for ch in ILLEGAL_FOLDER_CHARS:
        name = name.replace(ch, " ")
    return name


def generate_all(
    root: Path,
    dest_root: Path,
    info,
    session: word_writer.Session,
    package: dossier_packages.DossierPackage = None,
) -> None:
    """Tao ho so theo bo da chon (mac dinh la full) o thu muc tam local roi copy sang dest_root."""
    if package is None:
        package = dossier_packages.get_package("full")

    staging_root = Path(tempfile.mkdtemp(prefix="tao_ho_so_"))
    try:
        common_tokens = tokens.build_common_tokens(info)

        print(f"Dang sao chep file mau cho '{package.name}'...")
        copy_templates(root, staging_root, package=package)

        if package.copy_head_cv:
            print("Dang sao chep CV chu nhiem de tai...")
            copy_head_cv(root, staging_root, info)

        if package.copy_expert_cvs:
            print("Dang sao chep CV chuyen gia da khai...")
            copy_expert_cvs(root, staging_root, info, package=package)

        for section_id in package.sections:
            cfg = section_engine.load_section_config(section_id)
            section_dest = staging_root / cfg.output_dir
            if section_dest.exists():
                print(f"Dang sinh {cfg.name}...")
                section_engine.generate(session, section_dest, info, common_tokens, cfg)

        shutil.copytree(staging_root, dest_root, dirs_exist_ok=True)
    except Exception:
        print(f"  [LUU Y] Cac file da xu ly tam thoi con luu tai: {staging_root} (de kiem tra loi)")
        raise
    else:
        shutil.rmtree(staging_root, ignore_errors=True)




def main() -> None:
    root = paths.project_root()

    checklist_path = root / excel_reader.CHECKLIST_FILENAME
    sheet_name = None

    try:
        sheet_name, package = parse_cli_args(checklist_path)

        print(f"Dang doc du lieu tu Excel checklist (Sheet: '{sheet_name}')...")
        info = excel_reader.load_project_data(checklist_path, sheet_name)

        session = word_writer.Session()
        print(f"Che do ghi Word dang dung: {session.backend}")
        if session.backend == "docx":
            print(
                "  [LUU Y] Khong co Word COM - dung fallback python-docx thuan.\n"
                "  Mot vai cho xoa dong trong co the con sot dong trong, script se canh bao khi gap."
            )

        dest_dir_name = f"Hồ sơ - {_sanitize_folder_name(info.title)} ({info.year})"
        dest_root = root / dest_dir_name

        try:
            generate_all(root, dest_root, info, session, package=package)
        finally:
            session.quit()

        print(f"XONG. Bo ho so da tao tai: {dest_root}")


    except KeyError as e:
        # KeyError thuong xay ra khi SHEET_NAME sai ten, nhung cung co the la
        # ma muc (vd C09) bi thieu trong sheet dung - kiem tra ten sheet truoc
        # de dua ra thong bao dung trong tam.
        try:
            wb = openpyxl.load_workbook(root / excel_reader.CHECKLIST_FILENAME)
            sheet_names = wb.sheetnames
        except Exception:
            sheet_names = None

        if sheet_names is not None and sheet_name is not None and sheet_name not in sheet_names:
            print(f"[LOI] Khong tim thay sheet '{sheet_name}' trong file Excel checklist.")
            print("Cac sheet hien co trong file:")
            for name in sheet_names:
                print(f"  - {name}")
            print(
                "Neu ban truyen ten sheet qua dong lenh, hay kiem tra lai chinh ta "
                "(xem HUONG_DAN.md muc 7)."
            )
        else:
            print("Da xay ra loi khi tao ho so. Chi tiet loi:")
            print(f"  {e}")
            print("Vui long xem HUONG_DAN.md muc 7 de biet cach khac phuc.")
        sys.exit(1)

    except Exception as e:
        print("Da xay ra loi khi tao ho so. Chi tiet loi:")
        print(f"  {e}")
        print("Vui long xem HUONG_DAN.md muc 7 de biet cach khac phuc.")
        sys.exit(1)


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")
    main()
