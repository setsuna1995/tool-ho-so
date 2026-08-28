# migrate_add_tokens_sheet.py
from pathlib import Path

import openpyxl

import excel_reader

CHECKLIST_PATH = Path(__file__).resolve().parent / excel_reader.CHECKLIST_FILENAME
TOKENS_SHEET_NAME = "_Tokens"

HEADERS = ["token_name", "code", "kind", "param", "note"]

TOKEN_SPECS = [
    ("TEN_DE_TAI", "A01", "raw", "", "Tên đề tài"),
    ("NAM", "A03", "raw", "", "Năm thực hiện hồ sơ"),
    ("DON_VI_CHU_TRI", "A04", "raw", "", "Cơ quan chủ trì"),
    ("DON_VI_DOI_TAC", "A06", "raw_or_placeholder", "", "Cơ quan phối hợp"),
    ("CHU_NHIEM_HO_TEN", "B01", "person_ho_ten", "", "Chủ nhiệm đề tài - có học hàm/học vị"),
    ("CHU_NHIEM_TEN", "B01", "person_ten", "", "Chủ nhiệm đề tài - chỉ tên"),
    ("DONG_CHU_NHIEM_TEN", "B02", "person_ten", "", "Đồng chủ nhiệm đề tài - chỉ tên"),
    ("DONG_CHU_NHIEM_HO_TEN", "B02", "person_ho_ten", "", "Đồng chủ nhiệm đề tài - có học hàm/học vị"),
    ("THU_KY_DE_TAI", "B03", "person_ho_ten", "", "Thư ký đề tài"),
    ("THOI_GIAN_BAT_DAU", "A05", "timeline_start", "", "Mốc bắt đầu (MM/YYYY)"),
    ("THOI_GIAN_KET_THUC", "A05", "timeline_end", "", "Mốc kết thúc (MM/YYYY)"),
    ("DIA_DIEM_TRIEN_KHAI", "A07", "raw_or_placeholder", "……………………………", "Địa điểm triển khai nghiên cứu"),
]


def _write_tokens_sheet(wb) -> None:
    if TOKENS_SHEET_NAME in wb.sheetnames:
        del wb[TOKENS_SHEET_NAME]
    ws = wb.create_sheet(TOKENS_SHEET_NAME)
    ws.sheet_state = "hidden"
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for row_i, spec in enumerate(TOKEN_SPECS, start=2):
        for col_i, value in enumerate(spec, start=1):
            ws.cell(row=row_i, column=col_i, value=value)


def add_tokens_sheet(checklist_path: Path = CHECKLIST_PATH) -> None:
    wb = openpyxl.load_workbook(checklist_path)
    _write_tokens_sheet(wb)
    wb.save(checklist_path)


if __name__ == "__main__":
    add_tokens_sheet()
    print("Da tao/cap nhat sheet _Tokens voi 12 token mac dinh.")
