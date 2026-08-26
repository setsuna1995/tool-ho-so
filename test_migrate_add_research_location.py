from pathlib import Path

import openpyxl
import pytest

import migrate_add_research_location as migrate
import excel_reader

CHECKLIST_PATH = Path(__file__).resolve().parent / excel_reader.CHECKLIST_FILENAME
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]


def test_add_research_location_field_adds_a07_to_both_sheets():
    migrate.add_research_location_field()

    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        codes = [row[0].value for row in ws.iter_rows(min_row=5, max_col=1)]
        assert "A07" in codes


def test_add_research_location_field_is_idempotent():
    migrate.add_research_location_field()
    migrate.add_research_location_field()

    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        codes = [row[0].value for row in ws.iter_rows(min_row=5, max_col=1)]
        assert codes.count("A07") == 1


def _code_index(ws):
    index = {}
    for row in ws.iter_rows(min_row=5, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and not cell.value.startswith("SEC_"):
            index[cell.value] = cell.row
    return index


def _build_minimal_checklist_workbook(tmp_path):
    """A synthetic workbook shaped like the real one around the insertion
    point (SEC_A, A01-A06 each with a self-referencing status formula,
    SEC_B, B01 also self-referencing) - enough to exercise insert_rows()
    and _fix_status_formula_row_refs() without touching the live,
    already-migrated workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in migrate.SHEET_NAMES:
        ws = wb.create_sheet(sheet_name)
        row = 5
        for code in ("SEC_A", "A01", "A02", "A03", "A04", "A05", "A06", "SEC_B", "B01"):
            ws.cell(row=row, column=1, value=code)
            if not code.startswith("SEC_"):
                ws.cell(
                    row=row,
                    column=6,
                    value=f'=IF(ISBLANK(C{row}), "❌ CHƯA ĐIỀN (BÁO LỖI)", "✅ Xong")',
                )
            row += 1
    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    return path


@pytest.mark.parametrize("sheet_name", migrate.SHEET_NAMES)
def test_shifted_rows_reference_their_own_row(monkeypatch, tmp_path, sheet_name):
    fixture_path = _build_minimal_checklist_workbook(tmp_path)
    monkeypatch.setattr(migrate, "CHECKLIST_PATH", fixture_path)

    migrate.add_research_location_field()

    wb = openpyxl.load_workbook(fixture_path, data_only=False)
    ws = wb[sheet_name]
    index = _code_index(ws)
    b01_row = index["B01"]
    formula = ws.cell(row=b01_row, column=6).value
    assert f"C{b01_row}" in formula
    assert f"C{b01_row - 1}" not in formula
