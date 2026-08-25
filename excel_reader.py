import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl

CHECKLIST_FILENAME = "Form checklist hồ sơ dự án.xlsx"

TIMELINE_PATTERN = re.compile(r"(\d{2})/(\d{4}).*?(\d{2})/(\d{4})")


def parse_timeline(text: str) -> Tuple[str, str]:
    """Tách mốc bắt đầu/kết thúc dạng MM/YYYY từ nội dung mốc thời gian (A05)."""
    match = TIMELINE_PATTERN.search(text)
    if not match:
        raise ValueError(
            f"Không đọc được mốc thời gian nghiên cứu (A05) từ nội dung '{text}'. "
            "Vui lòng nhập đúng định dạng có 2 mốc MM/YYYY, ví dụ: "
            "'Tháng 01/2027 đến tháng 12/2027'."
        )
    start_month, start_year, end_month, end_year = match.groups()
    return f"{start_month}/{start_year}", f"{end_month}/{end_year}"


@dataclass
class Person:
    name: str
    degree: str = ""
    org: str = ""


@dataclass
class CommitteeData:
    chair: Person
    reviewers: List[Person] = field(default_factory=list)
    members: List[Person] = field(default_factory=list)
    secretaries: List[Person] = field(default_factory=list)


@dataclass
class ProjectInfo:
    title: str
    research_type: str
    year: int
    host_org: str
    partner_org: Optional[str]
    timeline: str
    head: Person
    co_head: Optional[Person]
    project_secretary: Optional[Person]
    researchers: List[Person]
    ethics_committee: CommitteeData
    proposal_committee: CommitteeData
    acceptance_committee: CommitteeData
    head_cv_filename: str


def _build_code_index(ws) -> dict:
    index = {}
    for row in ws.iter_rows(min_row=5, max_col=1):
        cell = row[0]
        code = cell.value
        if isinstance(code, str) and not code.startswith("SEC_"):
            index[code] = cell.row
    return index


def _cell_text(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()


def _read_person(ws, index: dict, code: str) -> Optional[Person]:
    row = index.get(code)
    if row is None:
        raise KeyError(f"Không tìm thấy mã mục '{code}' trong checklist")
    name = _cell_text(ws, row, 3)
    if not name:
        return None
    return Person(name=name, degree=_cell_text(ws, row, 4), org=_cell_text(ws, row, 5))


def _read_text(ws, index: dict, code: str) -> str:
    row = index.get(code)
    if row is None:
        raise KeyError(f"Không tìm thấy mã mục '{code}' trong checklist")
    return _cell_text(ws, row, 3)


def parse_committee(ws, index: dict, prefix: str) -> CommitteeData:
    chair = _read_person(ws, index, f"{prefix}01")
    if chair is None:
        raise ValueError(f"Chủ tịch hội đồng ({prefix}01) là bắt buộc nhưng đang trống")

    reviewers = []
    for code in (f"{prefix}02", f"{prefix}03", f"{prefix}06"):
        person = _read_person(ws, index, code)
        if person:
            reviewers.append(person)

    members = []
    for code in (f"{prefix}04", f"{prefix}05", f"{prefix}07", f"{prefix}08"):
        person = _read_person(ws, index, code)
        if person:
            members.append(person)

    secretaries = []
    for code in (f"{prefix}09", f"{prefix}10"):
        person = _read_person(ws, index, code)
        if person is None:
            raise ValueError(f"Thư ký hội đồng ({code}) là bắt buộc nhưng đang trống")
        secretaries.append(person)

    return CommitteeData(chair=chair, reviewers=reviewers, members=members, secretaries=secretaries)


def load_project_data(xlsx_path: Path, sheet_name: str) -> ProjectInfo:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    index = _build_code_index(ws)

    title = _read_text(ws, index, "A01")
    if not title:
        raise ValueError("Tên đề tài (A01) đang trống trong checklist")

    year_raw = ws.cell(row=index["A03"], column=3).value
    if year_raw is None:
        raise ValueError("Năm thực hiện hồ sơ (A03) đang trống trong checklist")

    head = _read_person(ws, index, "B01")
    if head is None:
        raise ValueError("Chủ nhiệm đề tài (B01) là bắt buộc nhưng đang trống")

    researchers = []
    for i in range(4, 21):
        code = f"B{i:02d}"
        if code not in index:
            continue
        person = _read_person(ws, index, code)
        if person:
            researchers.append(person)

    partner_org = _read_text(ws, index, "A06") if "A06" in index else ""

    head_cv_filename = _cell_text(ws, index["F01"], 5) if "F01" in index else ""
    if not head_cv_filename:
        raise ValueError("Tên file CV của chủ nhiệm đề tài (F01) là bắt buộc nhưng đang trống")

    return ProjectInfo(
        title=title,
        research_type=_read_text(ws, index, "A02"),
        year=int(year_raw),
        host_org=_read_text(ws, index, "A04"),
        partner_org=partner_org or None,
        timeline=_read_text(ws, index, "A05"),
        head=head,
        co_head=_read_person(ws, index, "B02"),
        project_secretary=_read_person(ws, index, "B03"),
        researchers=researchers,
        ethics_committee=parse_committee(ws, index, "C"),
        proposal_committee=parse_committee(ws, index, "D"),
        acceptance_committee=parse_committee(ws, index, "E"),
        head_cv_filename=head_cv_filename,
    )
