import openpyxl
import pytest

import excel_reader
import paths

CHECKLIST_PATH = paths.project_root() / excel_reader.CHECKLIST_FILENAME
SHEET_VIAM = "Đề tài - Bánh ăn dặm VIAM 2027"
SHEET_BLANK = "Đề tài - Mẫu trắng dự án mới"

_MINIMAL_ROWS = {
    "A01": ("Đề tài test", None, None),
    "A02": ("TVCT_ĐGHQ", None, None),
    "A03": (2027, None, None),
    "A04": ("Cơ quan", None, None),
    "A05": ("Tháng 01/2027 đến tháng 12/2027", None, None),
    "B01": ("Chủ nhiệm", "TS.", "Org"),
    "B02": (None, None, None),
    "B03": (None, None, None),
    "C01": ("Chủ tịch C", "", ""),
    "C02": (None, None, None),
    "C03": (None, None, None),
    "C04": (None, None, None),
    "C05": (None, None, None),
    "C06": (None, None, None),
    "C07": (None, None, None),
    "C08": (None, None, None),
    "C09": ("Thư ký C1", "", ""),
    "C10": ("Thư ký C2", "", ""),
    "D01": ("Chủ tịch D", "", ""),
    "D02": (None, None, None),
    "D03": (None, None, None),
    "D04": (None, None, None),
    "D05": (None, None, None),
    "D06": (None, None, None),
    "D07": (None, None, None),
    "D08": (None, None, None),
    "D09": ("Thư ký D1", "", ""),
    "D10": ("Thư ký D2", "", ""),
    "E01": ("Chủ tịch E", "", ""),
    "E02": (None, None, None),
    "E03": (None, None, None),
    "E04": (None, None, None),
    "E05": (None, None, None),
    "E06": (None, None, None),
    "E07": (None, None, None),
    "E08": (None, None, None),
    "E09": ("Thư ký E1", "", ""),
    "E10": ("Thư ký E2", "", ""),
    "F01": ("Chủ nhiệm", "Chủ nhiệm đề tài", "cv.docx"),
}


def _build_minimal_workbook(tmp_path, overrides=None):
    rows = dict(_MINIMAL_ROWS)
    if overrides:
        rows.update(overrides)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    row_num = 5
    for code, (c, d, e) in rows.items():
        ws.cell(row=row_num, column=1, value=code)
        ws.cell(row=row_num, column=3, value=c)
        ws.cell(row=row_num, column=4, value=d)
        ws.cell(row=row_num, column=5, value=e)
        row_num += 1

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


def test_load_project_data_title_and_year():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.title == "Tư vấn hiệu quả công thức sản phẩm Bánh ăn dặm VIAM"
    assert data.year == 2027


def test_secretary_org_differs_correctly_between_committees():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.ethics_committee.secretaries[0].org == "Trung tâm NCKH - Viện VIAM"
    assert data.proposal_committee.secretaries[0].org == "Trung tâm NCKH - Viện VIAM"
    assert data.acceptance_committee.secretaries[0].org == "Viện Y học Ứng dụng Việt Nam"


def test_optional_partner_org_is_none_when_blank():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.partner_org is None


def test_ethics_committee_has_required_counts():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    c = data.ethics_committee
    assert c.chair.name == "Nguyễn Công Khẩn"
    assert len(c.reviewers) == 2
    assert len(c.members) == 2
    assert len(c.secretaries) == 2


def test_head_and_researchers():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.head.name == "Trương Hồng Sơn"
    assert data.co_head is None
    names = [p.name for p in data.researchers]
    assert "Lê Việt Anh" in names
    assert "Lê Minh Khánh" in names


def test_blank_template_sheet_raises_on_missing_required_field():
    with pytest.raises(ValueError):
        excel_reader.load_project_data(CHECKLIST_PATH, SHEET_BLANK)


def test_parse_timeline_extracts_start_and_end_month_year():
    start, end = excel_reader.parse_timeline("Tháng 01/2027 đến tháng 12/2027")
    assert start == "01/2027"
    assert end == "12/2027"


def test_parse_timeline_ignores_surrounding_wording():
    start, end = excel_reader.parse_timeline("Từ 03/2027 đến 09/2028")
    assert start == "03/2027"
    assert end == "09/2028"


def test_parse_timeline_raises_when_two_dates_not_found():
    with pytest.raises(ValueError):
        excel_reader.parse_timeline("Chưa xác định")


def test_head_cv_filename_is_read_from_f01(tmp_path):
    path = _build_minimal_workbook(tmp_path)
    data = excel_reader.load_project_data(path, "Test")
    assert data.head_cv_filename == "cv.docx"


def test_missing_f01_cv_filename_raises_value_error(tmp_path):
    path = _build_minimal_workbook(tmp_path, overrides={"F01": ("Chủ nhiệm", "Chủ nhiệm đề tài", "")})
    with pytest.raises(ValueError):
        excel_reader.load_project_data(path, "Test")


def test_real_checklist_head_cv_filename():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.head_cv_filename
