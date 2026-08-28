import openpyxl

import capnhat_nhan_su as wiring


def _build_project_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    row_num = 5
    for code, name, degree, org in rows:
        ws.cell(row=row_num, column=1, value=code)
        ws.cell(row=row_num, column=3, value=name)
        ws.cell(row=row_num, column=4, value=degree)
        ws.cell(row=row_num, column=5, value=org)
        row_num += 1
    return ws


def _build_checklist(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_project_sheet(
        wb,
        "Đề tài - Bánh ăn dặm VIAM 2027",
        [("B01", "Trương Hồng Sơn", "TS.BS.", "VIAM"), ("C01", "Nguyễn Công Khẩn", "GS.TS.", "Hội đồng Đạo đức")],
    )
    _build_project_sheet(wb, "Đề tài - Mẫu trắng dự án mới", [])
    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    return path


def test_wire_person_dropdowns_adds_name_dropdown_on_person_rows(tmp_path):
    path = _build_checklist(tmp_path)

    wiring.wire_person_dropdowns(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Bánh ăn dặm VIAM 2027"]
    refs = {ref for dv in ws.data_validations.dataValidation for ref in str(dv.sqref).split()}
    assert "C5" in refs  # B01 row


def test_wire_person_dropdowns_writes_literal_degree_and_org_values(tmp_path):
    path = _build_checklist(tmp_path)

    wiring.wire_person_dropdowns(path)

    # data_only=True mo phong dung cach excel_reader.load_project_data doc file:
    # neu D/E la cong thuc thi se doc ra None -> mat hoc ham/don vi.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Đề tài - Bánh ăn dặm VIAM 2027"]
    assert ws.cell(row=5, column=4).value == "TS.BS."
    assert ws.cell(row=5, column=5).value == "VIAM"
    assert ws.cell(row=6, column=4).value == "GS.TS."
    assert ws.cell(row=6, column=5).value == "Hội đồng Đạo đức"


def test_wire_person_dropdowns_clears_stale_degree_and_org_when_name_blank(tmp_path):
    path = _build_checklist(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Bánh ăn dặm VIAM 2027"]
    # Dong nhan su co ma muc nhung chua chon ten, D/E con sot du lieu cu.
    ws.cell(row=7, column=1, value="C02")
    ws.cell(row=7, column=4, value="RÁC CŨ")
    ws.cell(row=7, column=5, value="ĐƠN VỊ CŨ")
    wb.save(path)

    wiring.wire_person_dropdowns(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Đề tài - Bánh ăn dặm VIAM 2027"]
    assert not ws.cell(row=7, column=4).value
    assert not ws.cell(row=7, column=5).value


def test_wire_person_dropdowns_is_idempotent_no_duplicate_validations(tmp_path):
    path = _build_checklist(tmp_path)

    wiring.wire_person_dropdowns(path)
    wiring.wire_person_dropdowns(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Bánh ăn dặm VIAM 2027"]
    refs_c5 = [dv for dv in ws.data_validations.dataValidation if "C5" in str(dv.sqref).split()]
    assert len(refs_c5) == 1


def test_wire_person_dropdowns_seeds_nhan_su_before_wiring(tmp_path):
    path = _build_checklist(tmp_path)

    wiring.wire_person_dropdowns(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[wiring.NHAN_SU_SHEET_NAME]
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "Trương Hồng Sơn" in names
    assert "Nguyễn Công Khẩn" in names
