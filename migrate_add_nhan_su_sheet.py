# migrate_add_nhan_su_sheet.py
from pathlib import Path

import openpyxl

import excel_reader

CHECKLIST_PATH = Path(__file__).resolve().parent / excel_reader.CHECKLIST_FILENAME
NHAN_SU_SHEET_NAME = "_NhanSu"
SHEET_NAMES = ["Đề tài - Bánh ăn dặm VIAM 2027", "Đề tài - Mẫu trắng dự án mới"]

HEADERS = ["ten", "hoc_ham_hoc_vi", "don_vi", "dia_chi", "sdt", "email", "cccd", "mst", "so_tk", "ngan_hang"]

PERSON_CODE_PREFIXES = ("B", "C", "D", "E")


def _build_code_index(ws) -> dict:
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str) and not row[0].value.startswith("SEC_")
    }


def _collect_existing_people(wb) -> dict:
    """Quet 2 sheet du an hien co, gom (ten -> hoc ham, don vi) de seed
    _NhanSu - tranh viec chuyen cot hoc ham/don vi sang cong thuc tra cuu
    lam mat du lieu da nhap cua cac du an cu."""
    people = {}
    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        index = _build_code_index(ws)
        for code, row in index.items():
            if code[0] not in PERSON_CODE_PREFIXES:
                continue
            name = ws.cell(row=row, column=3).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            if name in people:
                continue
            degree = ws.cell(row=row, column=4).value or ""
            org = ws.cell(row=row, column=5).value or ""
            people[name] = (degree, org)
    return people


def add_nhan_su_sheet(checklist_path: Path = CHECKLIST_PATH) -> None:
    wb = openpyxl.load_workbook(checklist_path)

    if NHAN_SU_SHEET_NAME in wb.sheetnames:
        ws = wb[NHAN_SU_SHEET_NAME]
        existing_names = {
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        }
    else:
        ws = wb.create_sheet(NHAN_SU_SHEET_NAME)
        ws.sheet_state = "hidden"
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col, value=header)
        existing_names = set()

    people = _collect_existing_people(wb)
    next_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    for name, (degree, org) in people.items():
        if name in existing_names:
            continue
        ws.cell(row=next_row, column=1, value=name)
        ws.cell(row=next_row, column=2, value=degree)
        ws.cell(row=next_row, column=3, value=org)
        next_row += 1
        existing_names.add(name)

    wb.save(checklist_path)


if __name__ == "__main__":
    add_nhan_su_sheet()
    print("Da tao/cap nhat sheet _NhanSu va seed du lieu tu 2 sheet du an hien co.")
