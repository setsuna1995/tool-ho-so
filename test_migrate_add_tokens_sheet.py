# test_migrate_add_tokens_sheet.py
import openpyxl
import pytest

import migrate_add_tokens_sheet as migrate


def test_add_tokens_sheet_creates_sheet_with_all_default_tokens(tmp_path):
    checklist_path = tmp_path / "checklist.xlsx"
    openpyxl.Workbook().save(checklist_path)

    migrate.add_tokens_sheet(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    assert migrate.TOKENS_SHEET_NAME in wb.sheetnames
    ws = wb[migrate.TOKENS_SHEET_NAME]
    assert ws.sheet_state == "hidden"
    token_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "DONG_CHU_NHIEM_HO_TEN" in token_names
    assert len(token_names) == len(migrate.TOKEN_SPECS)


def test_add_tokens_sheet_is_idempotent(tmp_path):
    checklist_path = tmp_path / "checklist.xlsx"
    openpyxl.Workbook().save(checklist_path)

    migrate.add_tokens_sheet(checklist_path)
    migrate.add_tokens_sheet(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb[migrate.TOKENS_SHEET_NAME]
    token_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert len(token_names) == len(migrate.TOKEN_SPECS)


def test_add_tokens_sheet_row_maps_kind_and_code_correctly(tmp_path):
    checklist_path = tmp_path / "checklist.xlsx"
    openpyxl.Workbook().save(checklist_path)

    migrate.add_tokens_sheet(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb[migrate.TOKENS_SHEET_NAME]
    rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    r = rows["DIA_DIEM_TRIEN_KHAI"]
    assert ws.cell(row=r, column=2).value == "A07"
    assert ws.cell(row=r, column=3).value == "raw_or_placeholder"
    assert ws.cell(row=r, column=4).value == "……………………………"


def test_add_tokens_sheet_row_maps_contact_person_token(tmp_path):
    checklist_path = tmp_path / "checklist.xlsx"
    openpyxl.Workbook().save(checklist_path)

    migrate.add_tokens_sheet(checklist_path)

    wb = openpyxl.load_workbook(checklist_path)
    ws = wb[migrate.TOKENS_SHEET_NAME]
    rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    r = rows["DAU_MOI_LIEN_HE"]
    assert ws.cell(row=r, column=2).value == "A08"
    assert ws.cell(row=r, column=3).value == "raw_or_placeholder"
    assert ws.cell(row=r, column=4).value == "……"

