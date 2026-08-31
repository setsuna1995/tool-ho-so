# test_token_rules.py
import openpyxl
import pytest

import paths
import token_rules
import excel_reader


def _code_index(ws):
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str) and not row[0].value.startswith("SEC_")
    }


def _build_workbook(tmp_path, token_rows, project_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đề tài - Test"
    row_num = 5
    for code, name, degree, org in project_rows:
        ws.cell(row=row_num, column=1, value=code)
        ws.cell(row=row_num, column=3, value=name)
        ws.cell(row=row_num, column=4, value=degree)
        ws.cell(row=row_num, column=5, value=org)
        row_num += 1

    tokens_ws = wb.create_sheet(token_rules.TOKENS_SHEET_NAME)
    tokens_ws.cell(row=1, column=1, value="token_name")
    for i, (name, code, kind, param) in enumerate(token_rows, start=2):
        tokens_ws.cell(row=i, column=1, value=name)
        tokens_ws.cell(row=i, column=2, value=code)
        tokens_ws.cell(row=i, column=3, value=kind)
        tokens_ws.cell(row=i, column=4, value=param)

    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    return path


def test_resolve_tokens_raw_kind(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[("TEN_DE_TAI", "A01", "raw", "")],
        project_rows=[("A01", "Đề tài mẫu", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{TEN_DE_TAI}}"] == "Đề tài mẫu"


def test_resolve_tokens_raw_or_placeholder_kind_uses_param_when_blank(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[("DIA_DIEM", "A07", "raw_or_placeholder", "……")],
        project_rows=[("A07", None, None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{DIA_DIEM}}"] == "……"


def test_resolve_tokens_raw_or_placeholder_kind_tolerates_absent_code(tmp_path):
    """Ban checklist cu chua co ma muc A07/A06 -> dung placeholder, khong KeyError."""
    path = _build_workbook(
        tmp_path,
        token_rows=[("DIA_DIEM", "A07", "raw_or_placeholder", "……")],
        project_rows=[("A01", "Đề tài mẫu", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    index = _code_index(ws)
    assert "A07" not in index
    result = token_rules.resolve_tokens(wb, ws, index)
    assert result["{{DIA_DIEM}}"] == "……"


def test_resolve_tokens_raw_or_placeholder_kind_empty_param_absent_code(tmp_path):
    """DON_VI_DOI_TAC (A06) co param rong -> tra ve chuoi rong, khong KeyError."""
    path = _build_workbook(
        tmp_path,
        token_rows=[("DON_VI_DOI_TAC", "A06", "raw_or_placeholder", "")],
        project_rows=[("A01", "Đề tài mẫu", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{DON_VI_DOI_TAC}}"] == ""


def test_resolve_tokens_person_ho_ten_kind_combines_degree_and_name(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[("CHU_NHIEM_HO_TEN", "B01", "person_ho_ten", "")],
        project_rows=[("B01", "Nguyễn Văn A", "TS.", "Viện ABC")],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{CHU_NHIEM_HO_TEN}}"] == "TS. Nguyễn Văn A"


def test_resolve_tokens_person_ho_ten_kind_blank_person_is_empty_string(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[("DONG_CHU_NHIEM_HO_TEN", "B02", "person_ho_ten", "")],
        project_rows=[("B02", None, None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{DONG_CHU_NHIEM_HO_TEN}}"] == ""


def test_resolve_tokens_person_ten_kind_is_bare_name(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[("CHU_NHIEM_TEN", "B01", "person_ten", "")],
        project_rows=[("B01", "Nguyễn Văn A", "TS.", "Viện ABC")],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{CHU_NHIEM_TEN}}"] == "Nguyễn Văn A"


def test_resolve_tokens_timeline_start_and_end_kinds(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[
            ("BAT_DAU", "A05", "timeline_start", ""),
            ("KET_THUC", "A05", "timeline_end", ""),
        ],
        project_rows=[("A05", "Tháng 01/2027 đến tháng 12/2027", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb, ws, _code_index(ws))
    assert result["{{BAT_DAU}}"] == "01/2027"
    assert result["{{KET_THUC}}"] == "12/2027"


def test_resolve_tokens_unknown_kind_raises_value_error(tmp_path):
    path = _build_workbook(
        tmp_path,
        token_rows=[("FOO", "A01", "not_a_real_kind", "")],
        project_rows=[("A01", "x", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    with pytest.raises(ValueError):
        token_rules.resolve_tokens(wb, ws, _code_index(ws))


def test_resolve_tokens_missing_tokens_sheet_returns_empty_dict(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đề tài - Test"
    ws.cell(row=5, column=1, value="A01")
    ws.cell(row=5, column=3, value="x")
    path = tmp_path / "no_tokens_sheet.xlsx"
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Đề tài - Test"]
    result = token_rules.resolve_tokens(wb2, ws2, _code_index(ws2))
    assert result == {}


CHECKLIST_PATH = paths.project_root() / excel_reader.CHECKLIST_FILENAME
SHEET_VIAM = "Đề tài - Bánh ăn dặm VIAM 2027"


def test_real_checklist_resolves_new_and_existing_tokens_correctly():
    import excel_reader as er

    data = er.load_project_data(CHECKLIST_PATH, SHEET_VIAM)

    # Gia tri LITERAL doc tay tu checklist that (A01/A03/A07) - khong suy ra
    # tu chinh `data`, de test khong tu chung minh chinh no.
    assert (
        data.common_tokens["{{TEN_DE_TAI}}"]
        == "Tư vấn hiệu quả công thức sản phẩm Bánh ăn dặm VIAM"
    )
    assert data.common_tokens["{{NAM}}"] == "2027"
    # A07 dang trong trong checklist that -> dung placeholder tu cot `param`.
    assert data.common_tokens["{{DIA_DIEM_TRIEN_KHAI}}"] == "……………………………"
    # A08 chua duoc dien trong checklist that -> dung placeholder tu cot `param`.
    assert data.common_tokens["{{DAU_MOI_LIEN_HE}}"] == "……"

    assert data.common_tokens["{{CHU_NHIEM_HO_TEN}}"] == f"{data.head.degree} {data.head.name}".strip()
    assert data.common_tokens["{{DONG_CHU_NHIEM_HO_TEN}}"] == (
        f"{data.co_head.degree} {data.co_head.name}".strip() if data.co_head else ""
    )
    assert data.common_tokens["{{THU_KY_DE_TAI}}"] == (
        f"{data.project_secretary.degree} {data.project_secretary.name}".strip()
        if data.project_secretary
        else ""
    )
