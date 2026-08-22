import openpyxl
import pytest

import migrate_add_partner_org as migrate


def _code_index(ws):
    index = {}
    for row in ws.iter_rows(min_row=5, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and not cell.value.startswith("SEC_"):
            index[cell.value] = cell.row
    return index


@pytest.fixture(scope="module", autouse=True)
def run_migration_once():
    migrate.add_partner_org_field()


@pytest.mark.parametrize("sheet_name", migrate.SHEET_NAMES)
def test_a06_row_added_with_correct_label(sheet_name):
    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH, data_only=False)
    ws = wb[sheet_name]
    index = _code_index(ws)
    assert "A06" in index
    a06_row = index["A06"]
    assert ws.cell(row=a06_row, column=2).value == "Cơ quan phối hợp (Tùy chọn)"
    assert f"C{a06_row}" in ws.cell(row=a06_row, column=6).value


@pytest.mark.parametrize("sheet_name", migrate.SHEET_NAMES)
def test_shifted_rows_reference_their_own_row(sheet_name):
    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH, data_only=False)
    ws = wb[sheet_name]
    index = _code_index(ws)
    b01_row = index["B01"]
    formula = ws.cell(row=b01_row, column=6).value
    assert f"C{b01_row}" in formula
    assert f"C{b01_row - 1}" not in formula
