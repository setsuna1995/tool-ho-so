# test_migrate_remove_cv_filename_column.py
import openpyxl

import migrate_remove_cv_filename_column as migrate


def _code_index(ws):
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str)
    }


def test_removes_e_column_validation_on_f_rows():
    migrate.remove_cv_filename_column()

    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH)
    for sheet_name in migrate.SHEET_NAMES:
        ws = wb[sheet_name]
        index = _code_index(ws)
        f01_row = index["F01"]
        all_refs = set()
        for dv in ws.data_validations.dataValidation:
            all_refs |= set(str(dv.sqref).split())
        assert f"E{f01_row}" not in all_refs


def test_f01_status_formula_no_longer_checks_role_or_filename_columns():
    migrate.remove_cv_filename_column()

    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]
    index = _code_index(ws)
    formula = ws.cell(row=index["F01"], column=6).value
    assert "ISBLANK(E" not in formula
    assert "ISBLANK(D" not in formula


def test_f04_status_formula_only_checks_name_column():
    migrate.remove_cv_filename_column()

    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]
    index = _code_index(ws)
    row = index["F04"]
    formula = ws.cell(row=row, column=6).value
    assert formula == f'=IF(ISBLANK(C{row}), "⚪ Tùy chọn (Trống)", "✅ Đã khai tên - CV sẽ tự khớp theo tên")'


def test_lists_sheet_removed():
    migrate.remove_cv_filename_column()

    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH)
    assert migrate.LISTS_SHEET not in wb.sheetnames


def test_is_idempotent():
    migrate.remove_cv_filename_column()
    migrate.remove_cv_filename_column()

    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH)
    ws = wb["Đề tài - Mẫu trắng dự án mới"]
    index = _code_index(ws)
    assert "F01" in index
