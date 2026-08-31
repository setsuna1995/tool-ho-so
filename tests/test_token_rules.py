# test_token_rules.py
import json
import openpyxl
import pytest

import paths
import token_rules
import excel_reader


def _code_index(ws):
    return {
        row[0].value: row[0].row
        for row in ws.iter_rows(min_row=5, max_col=1)
        if isinstance(row[0].value, str) and not row[0].value.startswith("SEC_")
    }


def _build_workbook_and_config(tmp_path, token_rows, project_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đề tài - Test"
    row_num = 5
    for code, name, degree, org in project_rows:
        ws.cell(row=row_num, column=1, value=code)
        ws.cell(row=row_num, column=3, value=name)
        ws.cell(row=row_num, column=4, value=degree)
        ws.cell(row=row_num, column=5, value=org)
        row_num += 1

    config_data = [
        {"token_name": name, "code": code, "kind": kind, "param": param, "note": ""}
        for name, code, kind, param in token_rows
    ]
    cfg_path = tmp_path / "config_tokens.json"
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(config_data, f)

    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    return path, cfg_path


def test_resolve_tokens_raw_kind(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("TEN_DE_TAI", "A01", "raw", "")],
        project_rows=[("A01", "Đề tài mẫu", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{TEN_DE_TAI}}"] == "Đề tài mẫu"


def test_resolve_tokens_supports_direct_token_in_column_a(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("DON_VI_CHU_TRI", "A04", "raw", "")],
        project_rows=[("{{DON_VI_CHU_TRI}}", "Viện Y học Ứng dụng", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    idx = excel_reader._build_code_index(ws)
    result = token_rules.resolve_tokens(ws, idx, config_path=cfg_path)
    assert result["{{DON_VI_CHU_TRI}}"] == "Viện Y học Ứng dụng"


def test_resolve_tokens_dynamic_custom_tokens_not_in_config(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("TEN_DE_TAI", "A01", "raw", "")],
        project_rows=[
            ("A01", "Đề tài mẫu", None, None),
            ("{{MA_SO_DE_TAI}}", "DT-2027-ABC", None, None),
            ("GHI_CHU_THEM", "Nghiên cứu độc quyền", None, None),
        ],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    idx = excel_reader._build_code_index(ws)
    result = token_rules.resolve_tokens(ws, idx, config_path=cfg_path)
    assert result["{{TEN_DE_TAI}}"] == "Đề tài mẫu"
    assert result["{{MA_SO_DE_TAI}}"] == "DT-2027-ABC"
    assert result["{{GHI_CHU_THEM}}"] == "Nghiên cứu độc quyền"



def test_resolve_tokens_raw_or_placeholder_kind_uses_param_when_blank(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("DIA_DIEM", "A07", "raw_or_placeholder", "……")],
        project_rows=[("A07", None, None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{DIA_DIEM}}"] == "……"


def test_resolve_tokens_raw_or_placeholder_kind_tolerates_absent_code(tmp_path):
    """Ban checklist cu chua co ma muc A07/A06 -> dung placeholder, khong KeyError."""
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("DIA_DIEM", "A07", "raw_or_placeholder", "……")],
        project_rows=[("A01", "Đề tài mẫu", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    index = _code_index(ws)
    assert "A07" not in index
    result = token_rules.resolve_tokens(ws, index, config_path=cfg_path)
    assert result["{{DIA_DIEM}}"] == "……"


def test_resolve_tokens_raw_or_placeholder_kind_empty_param_absent_code(tmp_path):
    """DON_VI_DOI_TAC (A06) co param rong -> tra ve chuoi rong, khong KeyError."""
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("DON_VI_DOI_TAC", "A06", "raw_or_placeholder", "")],
        project_rows=[("A01", "Đề tài mẫu", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{DON_VI_DOI_TAC}}"] == ""


def test_resolve_tokens_person_ho_ten_kind_combines_degree_and_name(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("CHU_NHIEM_HO_TEN", "B01", "person_ho_ten", "")],
        project_rows=[("B01", "Nguyễn Văn A", "TS.", "Viện ABC")],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{CHU_NHIEM_HO_TEN}}"] == "TS. Nguyễn Văn A"


def test_resolve_tokens_person_ho_ten_kind_blank_person_is_empty_string(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("DONG_CHU_NHIEM_HO_TEN", "B02", "person_ho_ten", "")],
        project_rows=[("B02", None, None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{DONG_CHU_NHIEM_HO_TEN}}"] == ""


def test_resolve_tokens_person_org_kind(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("ORG", "B01", "person_org", "")],
        project_rows=[("B01", "Nguyễn Văn A", "TS.BS.", "Viện ABC")],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result == {"{{ORG}}": "Viện ABC"}


def test_resolve_tokens_numbered_researchers_kind(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("RESEARCHERS", "B04", "numbered_researchers", "")],
        project_rows=[
            ("B04", "Lê Văn B", "ThS.", "Đơn vị X"),
            ("B05", "Trần Thị C", "BS.", "Đơn vị Y"),
        ],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result == {"{{RESEARCHERS}}": "1. ThS. Lê Văn B\n2. BS. Trần Thị C"}


def test_resolve_tokens_person_ten_kind_is_bare_name(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("CHU_NHIEM_TEN", "B01", "person_ten", "")],
        project_rows=[("B01", "Nguyễn Văn A", "TS.", "Viện ABC")],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{CHU_NHIEM_TEN}}"] == "Nguyễn Văn A"


def test_resolve_tokens_timeline_start_and_end_kinds(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[
            ("BAT_DAU", "A05", "timeline_start", ""),
            ("KET_THUC", "A05", "timeline_end", ""),
        ],
        project_rows=[("A05", "Tháng 01/2027 đến tháng 12/2027", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{BAT_DAU}}"] == "01/2027"
    assert result["{{KET_THUC}}"] == "12/2027"


def test_resolve_tokens_unknown_kind_raises_value_error(tmp_path):
    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[("FOO", "A01", "not_a_real_kind", "")],
        project_rows=[("A01", "x", None, None)],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    with pytest.raises(ValueError):
        token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)


def test_resolve_tokens_missing_config_returns_empty_dict(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đề tài - Test"
    ws.cell(row=5, column=1, value="A01")
    ws.cell(row=5, column=3, value="x")
    path = tmp_path / "checklist.xlsx"
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws2, _code_index(ws2), config_path=tmp_path / "nonexistent.json")
    assert result == {}


CHECKLIST_PATH = paths.project_root() / excel_reader.CHECKLIST_FILENAME
SHEET_VIAM = "Đề tài - Bánh ăn dặm VIAM 2027"


def test_real_checklist_resolves_new_and_existing_tokens_correctly():
    import excel_reader as er

    data = er.load_project_data(CHECKLIST_PATH, SHEET_VIAM)

    assert (
        data.common_tokens["{{TEN_DE_TAI}}"]
        == "Tư vấn hiệu quả công thức sản phẩm Bánh ăn dặm VIAM"
    )
    assert data.common_tokens["{{NAM}}"] == "2027"
    assert data.common_tokens["{{DIA_DIEM_TRIEN_KHAI}}"] == "……………………………"
    assert data.common_tokens["{{DAU_MOI_LIEN_HE}}"] == "……"

    assert data.common_tokens["{{CHU_NHIEM_HO_TEN}}"] == f"{data.head.degree} {data.head.name}".strip()
    assert data.common_tokens["{{CHU_NHIEM_DON_VI}}"] == data.head.org
    assert "1. Thạc sĩ Lê Việt Anh" in data.common_tokens["{{DANH_SACH_NGHIEN_CUU_VIEN}}"]
    assert data.common_tokens["{{CHU_TICH_HD_DAO_DUC}}"] == f"{data.ethics_committee.chair.degree} {data.ethics_committee.chair.name}".strip()
    assert data.common_tokens["{{CHU_TICH_HD_KHOA_HOC}}"] == f"{data.proposal_committee.chair.degree} {data.proposal_committee.chair.name}".strip()
    assert data.common_tokens["{{CHU_TICH_HD_NGHIEM_THU}}"] == f"{data.acceptance_committee.chair.degree} {data.acceptance_committee.chair.name}".strip()
    assert data.common_tokens["{{CHU_TICH_HD_NGHIEM_THU_TEN}}"] == data.acceptance_committee.chair.name
    assert data.common_tokens["{{DONG_CHU_NHIEM_HO_TEN}}"] == (
        f"{data.co_head.degree} {data.co_head.name}".strip() if data.co_head else ""
    )
    assert data.common_tokens["{{THU_KY_DE_TAI}}"] == (
        f"{data.project_secretary.degree} {data.project_secretary.name}".strip()
        if data.project_secretary
        else ""
    )


def test_build_template_context_creates_nested_and_flat_keys():
    import excel_reader as er

    data = er.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    context = token_rules.build_template_context(data)

    # Kiểm tra các trường phân cấp model
    assert context["project"]["title"] == "Tư vấn hiệu quả công thức sản phẩm Bánh ăn dặm VIAM"
    assert context["project"]["year"] == 2027
    assert context["project"]["head"]["name"] == data.head.name
    assert context["project"]["head"]["degree"] == data.head.degree
    assert context["project"]["head"]["full_name"] == f"{data.head.degree} {data.head.name}".strip()
    
    assert isinstance(context["researchers"], list)
    assert len(context["researchers"]) > 0
    assert "name" in context["researchers"][0]
    assert "full_name" in context["researchers"][0]

    assert "ethics" in context["committees"]
    assert context["committees"]["ethics"]["chair"]["name"] == data.ethics_committee.chair.name

    # Kiểm tra các trường token phẳng dùng chung
    assert context["TEN_DE_TAI"] == "Tư vấn hiệu quả công thức sản phẩm Bánh ăn dặm VIAM"
    assert context["CHU_NHIEM_HO_TEN"] == f"{data.head.degree} {data.head.name}".strip()


def test_resolve_tokens_date_vietnamese_kind(tmp_path):
    import datetime

    path, cfg_path = _build_workbook_and_config(
        tmp_path,
        token_rows=[
            ("NGAY_HOP", "G05", "date_vietnamese", "ngày …… tháng …… năm"),
            ("NGAY_TRONG", "G07", "date_vietnamese", "ngày …… tháng …… năm"),
        ],
        project_rows=[
            ("G05", "15/03/2027", None, None),
            ("G07", None, None, None),
        ],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Đề tài - Test"]
    result = token_rules.resolve_tokens(ws, _code_index(ws), config_path=cfg_path)
    assert result["{{NGAY_HOP}}"] == "ngày 15 tháng 03 năm 2027"
    assert result["{{NGAY_TRONG}}"] == "ngày …… tháng …… năm"


