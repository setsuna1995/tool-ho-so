import openpyxl
import pytest

import excel_reader
import paths

CHECKLIST_PATH = paths.project_root() / excel_reader.CHECKLIST_FILENAME
SHEET_VIAM = "Đề tài - Bánh ăn dặm VIAM 2027"
SHEET_BLANK = "Đề tài - Mẫu trắng dự án mới"

_MINIMAL_ROWS = {
    "A01": ("Đề tài test", None, None),
    "A02": ("TVCT_ĐGHQ", None, None),
    "A03": (2027, None, None),
    "A04": ("Cơ quan", None, None),
    "A05": ("Tháng 01/2027 đến tháng 12/2027", None, None),
    "B01": ("Chủ nhiệm", "TS.", "Org"),
    "B02": (None, None, None),
    "B03": (None, None, None),
    "C01": ("Chủ tịch C", "", ""),
    "C02": (None, None, None),
    "C03": (None, None, None),
    "C04": (None, None, None),
    "C05": (None, None, None),
    "C06": (None, None, None),
    "C07": (None, None, None),
    "C08": (None, None, None),
    "C09": ("Thư ký C1", "", ""),
    "C10": ("Thư ký C2", "", ""),
    "D01": ("Chủ tịch D", "", ""),
    "D02": (None, None, None),
    "D03": (None, None, None),
    "D04": (None, None, None),
    "D05": (None, None, None),
    "D06": (None, None, None),
    "D07": (None, None, None),
    "D08": (None, None, None),
    "D09": ("Thư ký D1", "", ""),
    "D10": ("Thư ký D2", "", ""),
    "E01": ("Chủ tịch E", "", ""),
    "E02": (None, None, None),
    "E03": (None, None, None),
    "E04": (None, None, None),
    "E05": (None, None, None),
    "E06": (None, None, None),
    "E07": (None, None, None),
    "E08": (None, None, None),
    "E09": ("Thư ký E1", "", ""),
    "E10": ("Thư ký E2", "", ""),
    "F01": ("Chủ nhiệm", "Chủ nhiệm đề tài", "cv.docx"),
}


def _build_minimal_workbook(tmp_path, overrides=None, replace_all=False):
    if replace_all and overrides is not None:
        rows = dict(overrides)
    else:
        rows = dict(_MINIMAL_ROWS)
        if overrides:
            rows.update(overrides)


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    row_num = 5
    for code, (c, d, e) in rows.items():
        ws.cell(row=row_num, column=1, value=code)
        ws.cell(row=row_num, column=3, value=c)
        ws.cell(row=row_num, column=4, value=d)
        ws.cell(row=row_num, column=5, value=e)
        row_num += 1

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


def test_load_project_data_title_and_year():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.title == "Tư vấn hiệu quả công thức sản phẩm Bánh ăn dặm VIAM"
    assert data.year == 2027


def test_secretary_org_differs_correctly_between_committees():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.ethics_committee.secretaries[0].org == "Trung tâm NCKH - Viện VIAM"
    assert data.proposal_committee.secretaries[0].org == "Trung tâm NCKH - Viện VIAM"
    assert data.acceptance_committee.secretaries[0].org == "Viện Y học Ứng dụng Việt Nam"


def test_optional_partner_org_is_none_when_blank():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.partner_org is None


def test_research_location_is_none_when_field_absent(tmp_path):
    path = _build_minimal_workbook(tmp_path)
    data = excel_reader.load_project_data(path, "Test")
    assert data.research_location is None


def test_research_location_is_read_when_present(tmp_path):
    path = _build_minimal_workbook(tmp_path, overrides={"A07": ("tỉnh Thái Bình", None, None)})
    data = excel_reader.load_project_data(path, "Test")
    assert data.research_location == "tỉnh Thái Bình"


def test_load_project_data_with_token_names_in_column_a(tmp_path):
    rows = dict(_MINIMAL_ROWS)
    del rows["A01"]
    del rows["A04"]
    rows["{{TEN_DE_TAI}}"] = ("Đề tài dùng Token Name", None, None)
    rows["{{DON_VI_CHU_TRI}}"] = ("Viện Y học Ứng dụng", None, None)
    path = _build_minimal_workbook(tmp_path, overrides=rows, replace_all=True)
    data = excel_reader.load_project_data(path, "Test")
    assert data.title == "Đề tài dùng Token Name"
    assert data.host_org == "Viện Y học Ứng dụng"





def test_ethics_committee_has_required_counts():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    c = data.ethics_committee
    assert c.chair.name == "Nguyễn Công Khẩn"
    assert len(c.reviewers) == 2
    assert len(c.members) == 2
    assert len(c.secretaries) == 2


def test_head_and_researchers():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert data.head.name == "Trương Hồng Sơn"
    assert data.co_head is None
    names = [p.name for p in data.researchers]
    assert "Lê Việt Anh" in names
    assert "Lê Minh Khánh" in names


def test_blank_template_sheet_raises_on_missing_required_field():
    with pytest.raises(ValueError):
        excel_reader.load_project_data(CHECKLIST_PATH, SHEET_BLANK)


def test_parse_timeline_extracts_start_and_end_month_year():
    start, end = excel_reader.parse_timeline("Tháng 01/2027 đến tháng 12/2027")
    assert start == "01/2027"
    assert end == "12/2027"


def test_parse_timeline_ignores_surrounding_wording():
    start, end = excel_reader.parse_timeline("Từ 03/2027 đến 09/2028")
    assert start == "03/2027"
    assert end == "09/2028"


def test_parse_timeline_raises_when_two_dates_not_found():
    with pytest.raises(ValueError):
        excel_reader.parse_timeline("Chưa xác định")


def test_read_expert_cvs_skips_blank_name_rows(tmp_path):
    path = _build_minimal_workbook(tmp_path, overrides={"F03": (None, None, None)})
    data = excel_reader.load_project_data(path, "Test")
    assert data.expert_cvs == []


def test_read_expert_cvs_reads_declared_rows_by_name(tmp_path):
    path = _build_minimal_workbook(
        tmp_path,
        overrides={
            "F03": ("Thư ký C", "Thư ký Đề tài", None),
            "F04": ("Chuyên gia D", "Ủy viên", None),
        },
    )
    data = excel_reader.load_project_data(path, "Test")
    codes = {e.code: e for e in data.expert_cvs}
    assert set(codes) == {"F03", "F04"}
    assert codes["F03"].name == "Thư ký C"
    assert codes["F03"].role == "Thư ký Đề tài"


def test_real_checklist_expert_cvs_returns_list():
    data = excel_reader.load_project_data(CHECKLIST_PATH, SHEET_VIAM)
    assert isinstance(data.expert_cvs, list)


def test_load_project_data_auto_lookups_nhan_su_when_degree_or_org_blank(tmp_path):
    # Tạo workbook có sheet _NhanSu và sheet Test với B01 chỉ có tên, degree/org để trống
    wb = openpyxl.Workbook()
    ws_test = wb.active
    ws_test.title = "Test"
    
    # Thiết lập _NhanSu
    ws_ns = wb.create_sheet("_NhanSu")
    ws_ns.cell(row=1, column=1, value="ten")
    ws_ns.cell(row=1, column=2, value="hoc_ham_hoc_vi")
    ws_ns.cell(row=1, column=3, value="don_vi")
    ws_ns.cell(row=2, column=1, value="Trương Hồng Sơn")
    ws_ns.cell(row=2, column=2, value="TS.BS.")
    ws_ns.cell(row=2, column=3, value="Viện Y học Ứng dụng")

    rows = dict(_MINIMAL_ROWS)
    rows["B01"] = ("Trương Hồng Sơn", "", "")  # Để trống degree/org
    
    row_num = 5
    for code, (c3, c4, c5) in rows.items():
        ws_test.cell(row=row_num, column=1, value=code)
        ws_test.cell(row=row_num, column=3, value=c3)
        ws_test.cell(row=row_num, column=4, value=c4)
        ws_test.cell(row=row_num, column=5, value=c5)
        row_num += 1

    path = tmp_path / "checklist.xlsx"
    wb.save(path)

    data = excel_reader.load_project_data(path, "Test")
    assert data.head.name == "Trương Hồng Sơn"
    assert data.head.degree == "TS.BS."
    assert data.head.org == "Viện Y học Ứng dụng"


