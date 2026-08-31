import openpyxl
import pytest

import migrate_remove_template_config_sheet as migrate


@pytest.fixture(scope="module", autouse=True)
def run_migration_once():
    migrate.remove_template_config_sheet()


def test_config_sheet_no_longer_present():
    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH, read_only=True)
    assert migrate.CONFIG_SHEET_NAME not in wb.sheetnames


def test_running_migration_twice_is_safe():
    migrate.remove_template_config_sheet()
    wb = openpyxl.load_workbook(migrate.CHECKLIST_PATH, read_only=True)
    assert migrate.CONFIG_SHEET_NAME not in wb.sheetnames
