# test_migrate_tokens_config.py
import json
import openpyxl
import pytest

import migrate_tokens_config as migrate


def test_save_config_tokens_creates_json_file(tmp_path):
    config_path = tmp_path / "config_tokens.json"
    migrate.save_config_tokens(config_path)

    assert config_path.exists()
    with open(config_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    assert len(data) == len(migrate.TOKEN_SPECS)
    token_names = [item["token_name"] for item in data]
    assert "TEN_DE_TAI" in token_names
    assert "CHU_NHIEM_HO_TEN" in token_names
    assert "DANH_SACH_NGHIEN_CUU_VIEN" in token_names


def test_remove_tokens_sheet_from_checklist_deletes_sheet_if_present(tmp_path):
    checklist_path = tmp_path / "checklist.xlsx"
    wb = openpyxl.Workbook()
    wb.create_sheet("_Tokens")
    wb.save(checklist_path)

    migrate.remove_tokens_sheet_from_checklist(checklist_path)

    wb2 = openpyxl.load_workbook(checklist_path)
    assert "_Tokens" not in wb2.sheetnames
